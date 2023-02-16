#-*- coding:utf-8 -*-
import yaml, json, base64, asyncio, os, sys
from nornir import InitNornir
from nornir.plugins import *
from nornir.plugins.inventory import *
from nornir.plugins.runners import *
from nornir.plugins.functions import *
from nornir.plugins.processors import *
from nornir.plugins.tasks import *
from nornir.plugins.connections import *
from nornir_netmiko.tasks import netmiko_send_config, netmiko_send_command
from nornir.core.task import Task, Result
from jinja2 import Template
from datetime import datetime, timedelta
from openpyxl import load_workbook
from lib.BlankNone import *
import pandas as pd
from operator import eq
import zipfile, collections, shutil, platform
from nornir.core.plugins.connections import ConnectionPluginRegister
from nornir_netmiko.connections import Netmiko
import re, ipaddress

###### pyinstaller 에서 netmiko 플러그인없이 빌드되는 경우가 있는듯
ConnectionPluginRegister.register("netmiko", Netmiko)


class ProcessAuto():
	def __init__(self, expired):
   
		###### 기간 제한 S #####
		now = datetime.now()		
		now = now.strftime("%Y%m%d")
		if int(expired) < int(now):
			raise Exception("oops!!")
		###### 기간 제한 E #####

		if eq(platform.system().lower(), "windows"):
			self.path = "./"
		else:
			self.path = os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]) + "/"

		self.db = self.path + "db/db.xlsx"
		self.inventory = self.path + "inventory.xlsx"
		self.lastConfigGen = "" ## 마지막으로 실행한 config 생성 session
		self.fullSession = {"fullv4":["init", "base", "loop0", "p2pip", "bgpv4", "etcport", "vxlan"] , "fullv6": ["init", "base", "loop0", "p2pipv6", "bgpv6", "etcport", "vxlan"]}
		self.sessions = list(set(self.fullSession["fullv4"] + self.fullSession["fullv6"]))
  
		# self.ymlInit()
		# self.norInit()
  
	def resource_path(self, relative_path):
		try:
			# PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수
			base_path = sys._MEIPASS
		except Exception:
			base_path = os.path.abspath(".")
		return os.path.join(base_path, relative_path)


	def generateInventory(self, inventory_file, excelVar):
		"""
		엑셀에서 데이터를 읽어 inventory 정보 처리
		d1.yml, all.yml 파일 생성
		toplogy 이미지 생성
		"""
		info = pd.read_excel(inventory_file, sheet_name="Var").fillna("")

		mgmtVrf = info.loc[[6], ["Variables Value"]].values[0][0]
		mgmtInterface = "Management1"
		mgmtGw = info.loc[[7],["Variables Value"]].values[0][0]
		macAging = info.loc[[16],["Variables Value"]].values[0][0]
		arpAging = info.loc[[17],["Variables Value"]].values[0][0]
		clockTimeZone = info.loc[[9],["Variables Value"]].values[0][0]
		adminName = info.loc[[11],["Variables Value"]].values[0][0]
		adminPassword = info.loc[[12],["Variables Value"]].values[0][0]
		admin_privilege = info.loc[[13],["Variables Value"]].values[0][0]
		spanningTreeMode = info.loc[[5],["Variables Value"]].values[0][0]
		terminalLength = info.loc[[1],["Variables Value"]].values[0][0]
		terminalWidth = info.loc[[2],["Variables Value"]].values[0][0]
		logginBuffered = info.loc[[3],["Variables Value"]].values[0][0]
		p2pSubnet = info.loc[[14],["Prefix & Define Value"]].values[0][0]
		spineBGPAsn = info.loc[[3],["Prefix & Define Value"]].values[0][0]
		spinePrefix = info.loc[[9],["Prefix & Define Value"]].values[0][0]
		leafPrefix = info.loc[[10],["Prefix & Define Value"]].values[0][0]

		##### port map 정보 로드 S #####
		sheetName = excelVar["pd"]["portMap"]["sheetName"]
		headerRow = excelVar["pd"]["portMap"]["header"]
		spineCol = excelVar["pd"]["portMap"]["spine"]
		spinePortCol = excelVar["pd"]["portMap"]["spinePort"]
		spineIpCol = excelVar["pd"]["portMap"]["spineIp"]
		leafCol = excelVar["pd"]["portMap"]["leaf"]
		leafPortCol = excelVar["pd"]["portMap"]["leafPort"]
		leafIpCol = excelVar["pd"]["portMap"]["leafIp"]

		switches = pd.read_excel(inventory_file, header=headerRow, sheet_name=sheetName)[[spineCol, spinePortCol, spineIpCol, leafCol, leafPortCol, leafIpCol]].dropna(axis=0)

		portMap = {}
		topologyInterfaces = {} ##### eve-ng, pnetlab 포톨로지파일 생성용 데이터

		for idx, switch in switches.iterrows():
			id = idx + 1
			###### 포트값에서 숫자 추출, 
			sPortID = re.sub(r'[^0-9]', '', switch[spinePortCol])
			ePortID = re.sub(r'[^0-9]', '', switch[leafPortCol])
			topologyInterfaces.setdefault(id, 
				{
					"START": switch[spineCol],
					"SPORT": switch[spinePortCol],
					"S_IP": switch[spineIpCol], 
					"S_ID": int(sPortID),
					"END": switch[leafCol],
					"EPORT": switch[leafPortCol],
					"E_IP": switch[leafIpCol],
					"E_ID": int(ePortID)
				}
			)
		
			##### P2P S #####  
			## spine switch 정리
			p = re.compile(spinePrefix)
			if (p.match(str(switch[spineCol]))):
				spine = switch[spineCol]
				if not spine in portMap:
					portMap.setdefault(
						spine,  {
							"INTERFACES": [{ "ETHERNET": switch[spinePortCol], "IP": switch[spineIpCol] }],
							"ETC_PORTS": { "IP": "", "INTERFACES": [] }
						}
					)
				else:
					portMap[spine]["INTERFACES"].append({ "ETHERNET": switch[spinePortCol], "IP": switch[spineIpCol] })
					
			## leaf switch 정리
			p = re.compile(leafPrefix)
			leaf = switch[spineCol]
			if (p.match(str(switch[leafCol])) and not p.match(leaf)):
				leaf = switch[leafCol]
				if not leaf in portMap:
					portMap.setdefault(
						leaf,  {
							"INTERFACES": [{ "ETHERNET": switch[leafPortCol], "IP": switch[leafIpCol] }],
							"ETC_PORTS": { "IP": "", "INTERFACES": [] }
						}
					)
				else:
					portMap[leaf]["INTERFACES"].append({ "ETHERNET": switch[leafPortCol], "IP": switch[leafIpCol] })
			##### P2P E ##### 
			
			##### port channell S #####
			p = re.compile(leafPrefix)
			# print(switch[spineCol], switch[leafCol])
			if (p.match(str(switch[leafCol])) and p.match(str(switch[spineCol]))):

				leaf = switch[leafCol]
				portMap[leaf]["ETC_PORTS"]["IP"] = switch[leafIpCol]
				portMap[leaf]["ETC_PORTS"]["INTERFACES"].append({ "ETHERNET": switch[leafPortCol] })

				leaf = switch[spineCol]
				portMap[leaf]["ETC_PORTS"]["IP"] = switch[spineIpCol]
				portMap[leaf]["ETC_PORTS"]["INTERFACES"].append({ "ETHERNET": switch[spinePortCol] })
		
			##### port channell E #####	
			
		# print(topologyInterfaces)
		##### eve-ng, pnetlab 포톨로지파일 생성용 데이터 생성 S ######
		with BlankNone(), open(self.path + "inventory/topologyInterfaces.yml", "w") as inv:
			inv.write(yaml.dump(topologyInterfaces, sort_keys=False))
		inv.close()
		##### eve-ng, pnetlab 포톨로지파일 생성용 데이터 생성 E ######
		
		##### port map 정보 로드 E #####

		## 기본변수 로드
		with open(self.resource_path("./excelEnvriment.json"), "r", encoding='utf8') as f:
			excelVar = json.load(f)
		f.close()
		
		sheetName = excelVar["pd"]["switchIpInfo"]["sheetName"]
		headerRow = excelVar["pd"]["switchIpInfo"]["header"]
		hostNameCol = excelVar["pd"]["switchIpInfo"]["hostName"]
		mgmtCol = excelVar["pd"]["switchIpInfo"]["mgmt"]
		loopback0Col = excelVar["pd"]["switchIpInfo"]["loopback0"]
		bgpAsnCol = excelVar["pd"]["switchIpInfo"]["bgpAsn"]
		typeCol = excelVar["pd"]["switchIpInfo"]["type"]
		idCol = excelVar["pd"]["switchIpInfo"]["id"]
		loop1Col = excelVar["pd"]["switchIpInfo"]["loopback1"]
		
	
		## Switch 정보 로드
		switches = pd.read_excel(inventory_file, header=headerRow, sheet_name=sheetName)[[hostNameCol, mgmtCol, loopback0Col, bgpAsnCol, typeCol, idCol, loop1Col]].fillna("")
		
		config = {"hosts": None}

		data = {}
		## spine, leaf, bl 개수 체크
		topologySwitches = { "spine": 0, "leaf": 0, "bl": 0 }

		for idx, switch in switches.iterrows():
			hostname = switch[hostNameCol]
			mgmt = switch[mgmtCol]
			loop0 = switch[loopback0Col]
			spinePrefix = excelVar["spine"]["prefix"]

			##### spine ip 체크 , leaf 일때 bgp underlay ip ??
			spines = []
			# print(topologyInterfaces)
			if switch[typeCol] != "Spine":
				for id in topologyInterfaces:
					p = re.compile(spinePrefix)
					if hostname == topologyInterfaces[id]["END"] and (p.match(str(topologyInterfaces[id]["START"]))):
						hn = topologyInterfaces[id]["START"]
						ip = topologyInterfaces[id]["S_IP"] if "/" not in topologyInterfaces[id]["S_IP"] else topologyInterfaces[id]["S_IP"].split("/")[0]
						spines.append({ "HOSTNAME": hn, "IP": ip })
			
			data.setdefault(
				hostname,  {
						"hostname": mgmt,
						"port": 22,
						"username": "admin",
						"password": "admin",
						"platform": "eos",
						"data": {
							"HOSTNAME": hostname,
							"HOST_IP": mgmt,
							"LOOPBACK0": loop0,
							"LOOPBACK1": str(switch[loop1Col]) + "/32" if switch[loop1Col] != "" else "",
							"PERMIT_IP": str(ipaddress.IPv4Interface(str(switch[loop1Col]) + "/24").network) if switch[loop1Col] != "" else "",					
							"INTERFACES": portMap[hostname]["INTERFACES"] if hostname in portMap else "",
							"ETC_PORTS": portMap[hostname]["ETC_PORTS"] if hostname in portMap else "",
							"BGP_ASN": int(switch[bgpAsnCol]) if switch[typeCol] == "Spine" else switch[bgpAsnCol],
							"SPINE_BGP_ASN": spineBGPAsn,
							"SPINES": spines,
							"ID": switch[idCol],
							"TYPE": switch[typeCol],
							"P2P_SUBNET": p2pSubnet
						}
				}
			)
		
			## spine, leaf, bl 갯수 체크
			v = topologySwitches[str(switch[typeCol]).lower()]
			topologySwitches[str(switch[typeCol]).lower()] = v + 1

		config = data

		with BlankNone(), open(self.path + "inventory/hosts.yml", "w") as inv:
			inv.write(yaml.dump(config, sort_keys=False))
		inv.close()
		
		# Group Vars all.yml 파일 생성
		data = {
			"TERMINAL_LENGTH": terminalLength,
			"TERMINAL_WIDTH": terminalWidth,
			"LOGGIN_BUFFERED": logginBuffered,
			"SPANNING_TREE_MODE": spanningTreeMode,
			"ADMIN_USER_NAME": adminName,
			"ADMIN_USER_PW": adminPassword,
			"CLOCK_TIMEZONE": clockTimeZone,
			"ARP_AGING": arpAging,
			"MAC_AGING": macAging,
			"ADMIN_PRIVILEGE": admin_privilege,
			"MGMT_VRF": mgmtVrf,
			"MGMT_INTERFACE": mgmtInterface,
			"MGMT_GW": mgmtGw,
			"BACKUP_FILENAME": "{{ inventory_hostname }}_{{ lookup('pipe', 'date +%Y%m%d%H%M%S') }}"
		}

		with open(self.path + 'inventory/templates/inventory/defaults.j2') as f:
			template = Template(f.read())
		f.close()
	
		with open(self.path + "inventory/defaults.yml", "w") as reqs:
			reqs.write(template.render(**data))
		reqs.close()

	def ymlInit(self):
		file_location = self.inventory

	
		with open(self.path + 'inventory/templates/inventory/config.j2') as f:
			template = Template(f.read())
		f.close()

		data = {
			"HOSTS_YML": self.path + "inventory/hosts.yml",
			"GROUPS_YML": self.path + "inventory/groups.yml",
			"DEFAULTS_YML": self.path + "inventory/defaults.yml",
		}

		with BlankNone(), open(self.path + "config.yml", "w") as inv:
			inv.write(template.render(**data))
		inv.close()

		with open(self.resource_path("excelEnvriment.json"), "r") as f:
			excelVar = json.load(f)
		f.close()

		self.generateInventory(file_location, excelVar)

  
  
	def norInit(self):
		self.nr = InitNornir(config_file=self.path + "config.yml")
		return self.nr
  

	def cleanConfigCall(self):
   
		##### 자동백업
		self.backupConfigCall(memo="스위치 초기화로인한 자동 백업")
  
		self.nr.run(task=self.cleanConfig)
		self.norInit()

		log = "초기화"
		self.log(desc=log, cfgBackUp=False)

		return self.nr
  
	def cleanConfig(self, task: Task):
		mgmtVrf = task.host.defaults.data["MGMT_VRF"]
		mgmtGw = task.host.defaults.data["MGMT_GW"]
		mgmtInterface = task.host.defaults.data["MGMT_INTERFACE"]
  
		vrfConfig = f"vrf {mgmtVrf}"
		vrfInstance = f"vrf instance {mgmtVrf}"


		cleanConfig = ["configure session", \
									"rollback clean-config", \
									f"{vrfInstance}", \
									f"interface {mgmtInterface}", \
									f"{vrfConfig}", \
									f"ip address {task.host.hostname}/24", \
									f"username {task.host.username} secret {task.host.password} privilege 15", \
									f"ip route {vrfConfig} 0/0 {mgmtGw}", \
									"ip routing", \
									"logging buffered 1000", \
									"logging console informational", \
									"logging monitor informational", \
									"logging synchronous level all", \
									"commit"]
  
		result = task.run(netmiko_send_config, config_commands=cleanConfig)
  
		# self.norInit()
		result = Result(
			host=task.host,
			result=task.run(netmiko_send_config, config_commands=cleanConfig)
		)
		return result
  
	def getBackupConfigList(self):
  
		workbook = load_workbook(filename=self.db, read_only=True, data_only=True)

		return workbook["db"]

	def selectedConfigCall(self, item):
		self.nr.data.reset_failed_hosts()
		cfg = f"config_backup/{item}"
		print("selected config call = ", cfg)
		self.nr.run(task=self.sendConfig, dir=cfg)
		return self.nr

			
	def sendConfig(self, task: Task, dir):
		"""
		config 배포
		"""
		cfg = self.path + f"inventory/{dir}/{task.host.name}.cfg"
		# print(cfg)
		result = Result(
			host=task.host,
			result=task.run(netmiko_send_config, config_file=cfg)
		)
		return result
 
	def sendConfigCall(self):
		"""
		config 배포 호출
		"""
		self.nr.data.reset_failed_hosts()
		self.nr.run(task=self.sendConfig, dir="config")
		session = self.lastConfigGen
		log = f'{session} config 배포'

		self.log(desc=log, cfgBackUp=True)
  
		return self.nr
   
	def sendConfigCheck(self):
		# print(f"마지막 config gen {self.lastConfigGen}")
		if self.lastConfigGen == "backup" or not self.lastConfigGen:
			return False
		else:
			return True
 
	def getLastConfigGen(self):
		return self.lastConfigGen
 
	def backupConfig(self, task: Task, now):
		
		taskHost = task.host
		taskResult = task.run(netmiko_send_command, command_string="show running-config")

		with open(self.path + f"inventory/config_backup/{now}/{taskHost}.cfg", "w") as inv:
			inv.write(taskResult[0].result)
			inv.close()
				
		return Result(
			host=taskHost,
			result=taskResult,
		)

	def backupConfigCall(self, memo):
		self.nr.data.reset_failed_hosts()
		now = datetime.now()
		folderName = now.strftime("%y%m%d%H%M%S")
		nowDate = now.strftime('%y-%m-%d %H:%M:%S')
  
		directory = self.path + f"inventory/config_backup"
		if not os.path.exists(directory):
			os.makedirs(directory)

		directory = self.path + f"inventory/config_backup/{folderName}"
		if not os.path.exists(directory):
			os.makedirs(directory)
  
		self.nr.run(task=self.backupConfig, now=folderName)
   
   
		workbook = load_workbook(filename=self.db, read_only=False, data_only=True)
		sheet = workbook["db"]
		sheet.append([memo, folderName, nowDate])
		workbook.save(self.db)
		workbook.close()
  
		return self.nr

	def createTopology(self, topology, cpu, ram, ethernetCount, version, net, icon, configInclude, leafHost):

		spineSwitchs = {}
		leafSwitchs = {}

		with open(self.path + "inventory/hosts.yml") as f:
			hosts = yaml.load(f, Loader=yaml.FullLoader)
		f.close()

		for host in hosts:
			hostType = hosts[host]["data"]["TYPE"]

			if (hostType.upper() == "SPINE"):
				spineSwitchs.setdefault(
					host, {}
				)
			else:
				leafSwitchs.setdefault(
					host, {}
				)
		
		spinesCount = len(spineSwitchs)
		# leafsCount = len(leafSwitchs)
		# width = 1920
		space = 200 ## switch icon 사이 간격
		leafSpace = 200
		spineTopMargin = 50
		leafTopMargin = 450
		hostTopMargin = 650
		# iconWidth = 65
  
		spineIdx = 0
		leafIdx = spinesCount
		
		##### 위치값 좌측위로 고정  
		s_start = 550
		l_start = 150
  
		configs = {}
		for host in hosts:
			hostType = hosts[host]["data"]["TYPE"]
			# print("host type = ", host)
			###### config 포함시 base64로 인코딩 처리
			if configInclude:
				# print("config 포함")
				with open(self.path + f"inventory/config/{host}.cfg") as r:
					config = base64.b64encode(r.read().encode('utf-8')).strip().decode('utf-8')
					configState = 1
				r.close()

			else:
				configState = 0
				config = ""
    
			if (hostType.upper() == "SPINE"):
				spineIdx = spineIdx + 1
				spineSwitchs[host]= {
															"ID": spineIdx,
															"LEFT": s_start + ((spineIdx-1) * space),
															"TOP": spineTopMargin,
															"CONFIG": configState
														}
				configs.setdefault(spineIdx, config)
			else:
				leafIdx = leafIdx + 1
				leafSwitchs[host]= {
															"ID": leafIdx,
															"LEFT": l_start + ((leafIdx-spinesCount-1) * leafSpace),
															"TOP": leafTopMargin,
															"CONFIG": configState
														}
				configs.setdefault(leafIdx, config)
  
		# print(configs)
		with open(self.path + "inventory/topologyInterfaces.yml") as f:
			topologyInterfaces = yaml.load(f, Loader=yaml.FullLoader)
			f.close()

		## leaf 수보다 ethernet interface 갯수가 작은경우
		leafSwitchCount = len(leafSwitchs)
		if leafSwitchCount > (int(ethernetCount) + 1):
			ethernetCount = leafSwitchCount + 1
   
		hostSwitchs = {}
  
		if leafHost:
			## leaf - host 포함인경우 1 더해줌
			ethernetCount = int(ethernetCount) + 1
			## leaf 스위치에서 남은 포트 확인
			leafHostInterfaces = {}
		
			for leaf in leafSwitchs:
				leafInterface = {f"Eth{x}": False for x in range(1, ethernetCount)}
				for interface in topologyInterfaces:
					if eq(topologyInterfaces[interface]["START"], leaf):
						leafInterface[topologyInterfaces[interface]["SPORT"]] = True
		
					if eq(topologyInterfaces[interface]["END"], leaf):
						leafInterface[topologyInterfaces[interface]["EPORT"]] = True
				
				leafHostInterfaces.setdefault(leaf, leafInterface)
			
			## 마지막 인터페이스 고유번호 확인
			lastInterfaceId = len(topologyInterfaces)
			## 호스트 만들기
			hostIdx = 0
			for leaf in leafSwitchs:
				for interface in leafHostInterfaces[leaf]:
					if not leafHostInterfaces[leaf][interface]:
						hostIdx = int(hostIdx) + 1
						host = f"HOST-{hostIdx}" 
						lastInterfaceId = int(lastInterfaceId) + 1
			
						## leaf ethernet 번호 추출
						leafEthId = list(re.findall("Eth(\d)", interface)[0])[0]

						topologyInterfaces.setdefault(lastInterfaceId, 
																					{
																						"START": leaf,
																						"SPORT": interface,
																						"S_IP": "",
																						"S_ID": leafEthId,
																						"END": host,
																						"EPORT": "Eth1",
																						"E_IP": "",
																						"E_ID": 1,
																					}
																					)

						hostSwitchs.setdefault(host, 
																		{
																		"ID": hostIdx + 100,
																		"LEFT": l_start + ((hostIdx-1) * leafSpace),
																		"TOP": hostTopMargin,
																		"CONFIG": 0
																		}
																	)
						break
    
		topologyName = topology
		data = {
			"NAME": topologyName,
			"EOS_VERSION": version,
			"CPU": cpu,
			"RAM": ram,
			"ETHERNET": ethernetCount,
			"SPINES": spineSwitchs,
			"LEAFS": leafSwitchs,
			"HOSTS": hostSwitchs,
			"NET": net,
			"INTERFACES": topologyInterfaces,
			"SWITCH_ICON": icon, 
			"CONFIGS": configs
		}
		
		directory = self.path + "topology"
		if not os.path.exists(directory):
			os.makedirs(directory)
  
		file = f"{topologyName}.unl"
  
		with open(self.path + 'inventory/templates/topology/topology.j2') as f:
			template = Template(f.read())
		f.close()

		with BlankNone(), open(f'{file}', "w", encoding='utf8') as reqs:   
			reqs.write(template.render(**data))
		reqs.close() 
    
    
		zipOutput = f'{directory}/{topologyName}.zip'
		
		zipFile = zipfile.ZipFile(zipOutput, "w")
		
		zipFile.write(file, compress_type=zipfile.ZIP_DEFLATED)

		zipFile.close()
  
		os.remove(file)
    
    
	def createSimpleTopology(self, topology, spinePrefix, leafPrefix, spinesCount, leafsCount, version, cpu, ram, ethernetCount, net, icon, leafHost):
    
		spineSwitchs = {}
		leafSwitchs = {}
  
		# width = 1920
		space = 200 ## switch icon 사이 간격
		leafSpace = 200
		spineTopMargin = 50
		leafTopMargin = 450
		hostTopMargin = 650
		# iconWidth = 50
  
		spineIdx = 0
		leafIdx = 0
  
		## lab은 mgmt 포함으로 이더넷 수를 계산하여 +1 를 추가한다.
		# ethernetCount = ethernetCount + 1

		###### ethernetCount 체크
		if (ethernetCount < leafsCount):
			ethernetCount = leafsCount + 3

		##### 위치값 좌측위로 고정  
		s_start = 550
		l_start = 150
     
		for i in range(1, spinesCount + 1):
			spineIdx = spineIdx + 1
			spineName = spinePrefix + str(i)
			spineSwitchs.setdefault(
					spineName, {
						"ID": spineIdx,
						"LEFT": s_start + (spineIdx * space),
						"TOP": spineTopMargin
					}
			)
   
   
		for i in range(1, leafsCount + 1):
			leafIdx = leafIdx + 1
			leafName = leafPrefix + str(i)
			leafSwitchs.setdefault(
					leafName, {
						"ID": leafIdx + spinesCount + 10,
						"LEFT": l_start + ((leafIdx-1) * leafSpace),
						"TOP": leafTopMargin
					}
			)

		topologyInterfaces = {}
  
		id = 0
		
		leafPort = 0
		for spine in spineSwitchs:
			leafPort = leafPort + 1
			spinePort = 0
			for leaf in leafSwitchs:
				id = id + 1
				spinePort = spinePort + 1
    
				topologyInterfaces.setdefault(
					id, {
						"START": spine,
						"SPORT": "Eth" + str(spinePort),
						"S_ID": int(spinePort),
						"END": leaf,
						"EPORT": "Eth" + str(leafPort),
						"E_ID": int(leafPort),
					}
				)
  

		## leaf 수보다 ethernet interface 갯수가 작은경우
		leafSwitchCount = len(leafSwitchs)
		if leafSwitchCount > (ethernetCount + 1):
			ethernetCount = leafSwitchCount + 1
   
		hostSwitchs = {}
		## leaf - host 포함인경우 1 더해줌
		if leafHost:
			ethernetCount = ethernetCount + 1
   
			## leaf 스위치에서 남은 포트 확인
			leafHostInterfaces = {}
   
			for leaf in leafSwitchs:
				leafInterface = {f"Eth{x}": False for x in range(1, ethernetCount)}
				for interface in topologyInterfaces:
					if eq(topologyInterfaces[interface]["END"], leaf):
						leafInterface[topologyInterfaces[interface]["EPORT"]] = True
				
				leafHostInterfaces.setdefault(leaf, leafInterface)
			## 마지막 인터페이스 고유번호 확인
			lastInterfaceId = len(topologyInterfaces)
			## 호스트 만들기
			hostIdx = 0
			for leaf in leafSwitchs:
				for interface in leafHostInterfaces[leaf]:
					if not leafHostInterfaces[leaf][interface]:
						hostIdx = int(hostIdx) + 1
						host = f"HOST-{hostIdx}" 
						lastInterfaceId = int(lastInterfaceId) + 1
			
						## leaf ethernet 번호 추출
						leafEthId = list(re.findall("Eth(\d)", interface)[0])[0]

						topologyInterfaces.setdefault(lastInterfaceId, 
																					{
																						"START": leaf,
																						"SPORT": interface,
																						"S_IP": "",
																						"S_ID": leafEthId,
																						"END": host,
																						"EPORT": "Eth1",
																						"E_IP": "",
																						"E_ID": 1,
																					}
																					)

						hostSwitchs.setdefault(host, 
																		{
																		"ID": hostIdx + 100,
																		"LEFT": l_start + ((hostIdx-1) * leafSpace),
																		"TOP": hostTopMargin,
																		"CONFIG": 0
																		}
																	)
						break
   
		topologyName = topology
  
		data = {
			"NAME": topologyName,
			"EOS_VERSION": version,
			"CPU": cpu,
			"RAM": ram,
			"ETHERNET": ethernetCount,
			"SPINES": spineSwitchs,
			"LEAFS": leafSwitchs,
			"HOSTS": hostSwitchs,
			"INTERFACES": topologyInterfaces,
			"NET": net,
			"SWITCH_ICON": icon
		}
		
		# print(data)
  
		directory = self.path + f"topology"
		if not os.path.exists(directory):
			os.makedirs(directory)
  
		file = f'{topologyName}.unl'
  
		with open(self.path + 'inventory/templates/topology/topology.j2') as f:
			template = Template(f.read())
		f.close()

		with BlankNone(), open(f'{file}', "w", encoding='utf8') as reqs:   
				reqs.write(template.render(**data))
		reqs.close() 
    
		zipOutput = f'{directory}/{topologyName}.zip'
		
		zipFile = zipfile.ZipFile(zipOutput, "w")
		
		zipFile.write(file, compress_type=zipfile.ZIP_DEFLATED)

		zipFile.close()
  
		os.remove(file)

	def createConfig(self, session):
		self.lastConfigGen = session
  
		workbook = load_workbook(filename=self.inventory, read_only=True, data_only=True)
		j2 = ""
  
		if session in self.fullSession:
			for ss in self.fullSession[session]:
				sheet = workbook[ss]
				if sheet.iter_rows():
					for row in sheet.iter_rows():
						j2 = j2 + row[0].value + "\n"
		else:
			sheet = workbook[session]
			if sheet.iter_rows():
				for row in sheet.iter_rows():
					j2 = j2 + row[0].value + "\n"
				
		workbook._archive.close()
  
		template = Template(j2)
      
		for host in self.nr.inventory.hosts:
			data = {}
			for k in self.nr.inventory.hosts[host].keys():
				data.setdefault(k, self.nr.inventory.hosts[host][k])

			with open(self.path + "inventory/config/" + host + ".cfg", "w", encoding='utf8') as reqs:   
				reqs.write(template.render(**data))
				reqs.close() 
		return self.nr

	def statusCheck(self, task: Task):

		data = {}
		result = result=task.run(netmiko_send_command, command_string="show clock | json")
  
		r = json.loads(result[0].result)
		data.setdefault(
			"timezone", r["timezone"]
		)
		year = r["localTime"]["year"] 
		month = r["localTime"]["month"] 
		day = r["localTime"]["dayOfMonth"] 
		hour = r["localTime"]["hour"] 
		min = r["localTime"]["min"] 
		sec = r["localTime"]["sec"] 
		now = f'{year}-{month}-{day} {hour}:{min}:{sec}' 
		data.setdefault(
			"date", now
		)
  
		result = task.run(netmiko_send_command, command_string="show version | json")

		r = json.loads(result[0].result)
		
		data.setdefault(
			"model", r["modelName"]
		)
  
		data.setdefault(
			"version", r["version"]
		)
		return data

	def statusCheckCall(self):
		self.nr.data.reset_failed_hosts()
		return self.nr.run(task=self.statusCheck)
		
  
	def pingTest(self, task: Task):
		self.nr.data.reset_failed_hosts()
		cmd = "show ip interface | json"
  
		dictPing = collections.OrderedDict()
		listPing = ["Name","Local_address","Peer_address","Result", "Mtu","Description","Vrf"]

		returnInfo = task.run(netmiko_send_command, command_string=cmd)
		returnInfo = json.loads(returnInfo[0].result)
		# print(returnInfo)
		# returnInfo = returnInfo.result[0]
		# print("==================================s")
		# print(returnInfo["interfaces"])
		# for item in returnInfo:
				# print(item)
		# print("==================================e")
		# pingTarget = []
		p = re.compile("Ethernet")
		
		for interface in returnInfo["interfaces"]:
			# print("interface = ", interface)
			#### ethernet interface만 체크
			if (p.match(interface)):
				maskLen = returnInfo["interfaces"][interface]["interfaceAddress"]["primaryIp"]["maskLen"]
				# print(maskLen)
				if not interface in dictPing.keys():
					dictPing[interface] = collections.OrderedDict(zip(listPing, [''] * len(listPing)))
					dictPing[interface]["vrf"] = returnInfo["interfaces"][interface]["vrf"]
					# line protocol status 를 view 에서 제거 하기 위해서
					if returnInfo["interfaces"][interface]["lineProtocolStatus"] != "up":
						dictPing[interface]["result"] = "down"
					else:
						dictPing[interface]["result"] = "up"
					
					interfacename = returnInfo["interfaces"][interface]["name"]
					dictPing[interface]["name"] = interfacename
					# peer_address 계산
					localAddr = returnInfo["interfaces"][interface]["interfaceAddress"]["primaryIp"]["address"]
					tempAddr = localAddr.split('.')
					if maskLen == 31:
						if int(tempAddr[3]) % 2 == 0:
							tempAddr[3] = str(int(tempAddr[3]) + 1)
						else:
							tempAddr[3] = str(int(tempAddr[3]) - 1)
					elif maskLen == 30:
						
						if int(tempAddr[3]) % 2 == 0:
							tempAddr[3] = str(int(tempAddr[3]) - 1)
						else:
							tempAddr[3] = str(int(tempAddr[3]) + 1)
					else:
						tempAddr = returnInfo["interfaces"][interface]["interfaceAddress"]["broadcastAddress"].split('.')
					dictPing[interface]["peerAddress"] = ".".join(tempAddr)
					dictPing[interface]["localAddress"] = localAddr + "/" + str(maskLen)
		# pprint.pprint(dictPing)
  
		pingResult = {"success":0, "fail": 0, "down": 0}
		for peer in dictPing:
			peerIP = dictPing[peer]["peerAddress"]
			source = dictPing[peer]["name"]
			vrf = dictPing[peer]["vrf"]
   
			cmd = f'ping vrf {vrf} {peerIP} source {source} repeat 1'
			print(f'{task.host} {cmd}')
			result = task.run(netmiko_send_command, command_string=cmd)
			if " 0% packet loss" in result[0].result:
				pingResult["success"] = int(pingResult["success"]) + 1
			else:
				pingResult["fail"] = int(pingResult["fail"]) + 1
    
			if dictPing[peer]["result"] == "down":
				pingResult["down"] = int(pingResult["down"]) + 1
		
		return pingResult
    
	def pingTestCall(self):
		self.nr.data.reset_failed_hosts()
		return self.nr.run(self.pingTest)

	def log(self, desc=None, cfgBackUp=False):
		directory = f"./log"
  
		now = datetime.now()
		logTime = now.strftime("%y%m%d%H%M%S")
		logFile = now.strftime("%y%m%d%H")
		folder = now.strftime("%y%m%d")
		nowDate = now.strftime('%y-%m-%d %H:%M:%S')
		logDateFolder = f"{directory}/{folder}"
		logCfg = f'{logDateFolder}/{logTime}'
  
		if not os.path.exists(directory):
			os.makedirs(directory)

		if not os.path.exists(logDateFolder):
			os.makedirs(logDateFolder)
   
		with open(f'{logDateFolder}/{logFile}', "a", encoding='utf8') as reqs:   
			for host in self.nr.inventory.hosts:
				if host in self.nr.data.failed_hosts:
					log = f'{nowDate} {host} {desc} 실패'
				else:
					log = f'{nowDate} {host} {desc} 완료'

				reqs.write(f'{log}\n')
		reqs.close()
		
		

		if cfgBackUp:
			if not os.path.exists(logCfg):
				shutil.copytree(self.path + "inventory/config/", f'{logCfg}')

	def getCmdResult(self, task:Task, cmd):
		result = task.run(netmiko_send_command, command_string=cmd)
		# print(result[0].result)
		return result

	def getCmdResultCall(self, cmd):
		result = self.nr.run(task=self.getCmdResult, cmd=cmd)
  
		data = {}

		for r in result:
			data.setdefault(r, result[r].result[0].result)

		# print(data)
		return data				
    
    