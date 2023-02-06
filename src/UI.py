import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import *
import yaml, json
import asyncio, os, sys
from nornir import InitNornir
from nornir.plugins import *
from nornir.plugins.inventory import *
from nornir.plugins.runners import *
from nornir.plugins.functions import *
from nornir.plugins.processors import *
from nornir.plugins.tasks import *
from nornir.plugins.connections import *
from nornir_netmiko.tasks import netmiko_send_config, netmiko_send_command
from nornir.core.task import Task, Result
from jinja2 import Template
from datetime import datetime
from openpyxl import load_workbook
from generators.BlankNone import *
from generators.generateInventory import generateInventory 
from operator import eq
import zipfile
from src.ProcessAuto import ProcessAuto
from src.ProcessLab import ProcessLab
from tkinter.messagebox import askyesno
import platform

class UI():
	def __init__(self, asyncio=asyncio, processAuto=ProcessAuto, processLab=ProcessLab):
		self.asyncio = asyncio
		self.processAuto = processAuto
		self.processLab = processLab
		self.loop = self.asyncio.get_event_loop()
		# if platform.system()=='Windows':
		# 	asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
		self.root = tk.Tk()
		self.root.title("i-Cloud")
		self.root.iconbitmap(self.resource_path("icloud.ico"))
		self.root.geometry("750x800+600+100") ## w, h, x, y
		self.root.resizable(False, False)
  
		if eq(platform.system().lower(), "windows"):
			self.path = "./"
			self.defaultButtonW = 18
			self.defaultTextBoxW = 40
			self.defaultComboBoxW = 38
			self.defaultbackupTextBoxW = 50
		else:
			self.path = os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]) + "/"
			self.defaultButtonW = 20
			self.defaultTextBoxW = 26
			self.defaultComboBoxW = 24
			self.defaultbackupTextBoxW = 36

		self.db = "./db/db.xlsx"
		self.inventory = "./inventory.xlsx"
		self.cpuSize = [1,2]
		self.ethernetSize = []
		for i in range(9, 97):
			self.ethernetSize.append(i)
   
		self.spineSize = [2,4]
		self.leafSize = []
		for i in range(2, 60):
			self.leafSize.append(i)
   
		self.ramSize = [1024, 2048, 3072, 4096]
		self.switchIcon = ["AristaSW.png", "c_switch_blue.png", "Switch Blue.png", "Router Blue.png", "Router.png", "NI2_L3switch.png", "NI2_Switch.png"]
		self.lastConfigGen = "" ## 마지막으로 실행한 config 생성 session
		self.fullSession = {"fullv4":["init", "base", "loop0", "p2pip", "bgpv4", "etcport", "vxlan", "fullv4"] , "fullv6": ["init", "base", "loop0", "p2pipv6", "bgpv6", "etcport", "vxlan", "fullv6"]}
		self.sessions = list(set(self.fullSession["fullv4"] + self.fullSession["fullv6"]))
		self.vendors = ["arista"]
  
		self.ymlInit()
		self.norInit()
  
  
		##### grid S #####
		self.frame_top = ttk.Frame(self.root)
		self.frame_top.pack(side="top")


		self.treeview=ttk.Treeview(self.frame_top, columns=["one", "two", "three"])
		self.treeview.tag_configure("green",foreground='green')
		self.treeview.tag_configure("red",foreground='red')

		self.treeview.column("#1", width=180, anchor="center")
		self.treeview.heading("one", text="name", anchor="center")

		self.treeview.column("#2", width=330, anchor="w")
		self.treeview.heading("two", text="description", anchor="center")
  
		self.treeview.column("#3", width=120, anchor="w")
		self.treeview.heading("three", text="datetime", anchor="center")


		self.treeview["show"] = "headings"
		self.treeview.pack()
		##### grid E #####
  
		##### 상태 체크 S #####
		self.statusFrame = ttk.Frame(self.root)
		self.statusFrame.pack(side="top")

		self.buttonStatus = tk.Button(self.statusFrame, text="상태체크", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.statusCheckCall()))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		##### 상태 체크 E #####

		##### tab 구현 S #####
		self.notebook = ttk.Notebook(self.root, width=620, height=500, )
		self.notebook.pack()
  
	
		##### ipv4
		self.ipv4Frame = ttk.Frame(self.root)
		self.notebook.add(self.ipv4Frame, text=" IPV4 ")

		self.buttonStatus = tk.Button(self.ipv4Frame, text="init Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("init")))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.ipv4Frame, text="base Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("base")))
		self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.ipv4Frame, text="loop0 Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("loop0")))
		self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.ipv4Frame, text="etcport Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("etcport")))
		self.buttonList.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.ipv4Frame, text="vxlan Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("vxlan")))
		self.buttonList.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonStatus = tk.Button(self.ipv4Frame, text="p2p IPv4 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("p2pip")))
		self.buttonStatus.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.ipv4Frame, text="BGP IPv4 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("bgpv4")))
		self.buttonBackUp.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
	
		self.buttonStatus = tk.Button(self.ipv4Frame, text="full IPv4 Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("fullv4")))
		self.buttonStatus.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.ipv4Frame, text="Config 배포", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.sendConfigCall()))
		self.buttonSelectedDeploy.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### ipv6
		self.ipv6Frame = ttk.Frame(self.root)
		self.notebook.add(self.ipv6Frame, text=" IPV6 ")

		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="init Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("init")))
		self.buttonStatusIpv6.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUpIpv6 = tk.Button(self.ipv6Frame, text="base Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("base")))
		self.buttonBackUpIpv6.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="loop0 Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("loop0")))
		self.buttonListIpv6.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="etcport Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("etcport")))
		self.buttonListIpv6.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="vxlan Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("vxlan")))
		self.buttonListIpv6.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="p2p IPv6 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("p2pipv6")))
		self.buttonStatusIpv6.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUpIpv6 = tk.Button(self.ipv6Frame, text="BGP IPv6 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("bgpv6")))
		self.buttonBackUpIpv6.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="full IPv6 Config 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createConfig("fullv6")))
		self.buttonStatusIpv6.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		self.buttonSelectedDeployIpv6 = tk.Button(self.ipv6Frame, text="Config 배포", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.sendConfigCall()))
		self.buttonSelectedDeployIpv6.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		##### topology
		self.topologyFrame = ttk.Frame(self.root)
		self.notebook.add(self.topologyFrame, text=" LAB-TOPOLOGY ")
  
  
		self.topologyLabel = ttk.Label(self.topologyFrame, text="이름(영문) : ")
		self.topologyLabel.grid(row=0, column=1, padx=(8, 8), pady=(8, 8))
		self.topology = StringVar()
		self.topologyTextBox = ttk.Entry(self.topologyFrame, width=self.defaultTextBoxW, textvariable=self.topology)
		self.topologyTextBox.grid(row=0, column=2, padx=(8, 8), pady=(8, 8))
  
		self.versionLabel = ttk.Label(self.topologyFrame, text="버전 : ")
		self.versionLabel.grid(row=1, column=1, padx=(8, 8), pady=(8, 8))
		self.version = StringVar()
		self.version.set("veos-4.27.4.1M-noztp")
		self.versionTextBox = ttk.Entry(self.topologyFrame, width=self.defaultTextBoxW, textvariable=self.version)
		self.versionTextBox.grid(row=1, column=2, padx=(8, 8), pady=(8, 8))

		self.ethernetLabel = ttk.Label(self.topologyFrame, text="Ethernet 개수 : ")
		self.ethernetLabel.grid(row=2, column=1, padx=(8, 8), pady=(8, 8))
		# self.ethernet = StringVar()
		self.ethernetComboBox = ttk.Combobox(self.topologyFrame, width=self.defaultComboBoxW, values=self.ethernetSize, state="readonly")
		self.ethernetComboBox.current(0)
		self.ethernetComboBox.grid(row=2, column=2, padx=(8, 8), pady=(8, 8))
  
		self.cpuLabel = ttk.Label(self.topologyFrame, text="CPU 개수 : ")
		self.cpuLabel.grid(row=3, column=1)
		# self.cpu = StringVar()
		self.cpuComboBox = ttk.Combobox(self.topologyFrame, width=self.defaultComboBoxW, values=self.cpuSize, state="readonly")
		self.cpuComboBox.current(1)
		self.cpuComboBox.grid(row=3, column=2, padx=(8, 8), pady=(8, 8))
  
		
		self.ramLabel = ttk.Label(self.topologyFrame, text="RAM 크기 : ")
		self.ramLabel.grid(row=4, column=1, padx=(8, 8), pady=(8, 8))
		# self.ram = StringVar()
		self.ramComboBox = ttk.Combobox(self.topologyFrame, width=self.defaultComboBoxW, values=self.ramSize, state="readonly")
		self.ramComboBox.current(3)
		self.ramComboBox.grid(row=4, column=2, padx=(8, 8), pady=(8, 8))
  
		self.cloudLabel = ttk.Label(self.topologyFrame, text="Cloud1 네트워크 생성 : ")
		self.cloudLabel.grid(row=5, column=1, padx=(8, 8), pady=(8, 8))
		self.cloud = IntVar()
		self.cloudCheckBox = ttk.Checkbutton(self.topologyFrame, width=self.defaultComboBoxW, variable=self.cloud)
		self.cloudCheckBox.grid(row=5, column=2, padx=(8, 8), pady=(8, 8))
  
		self.configIncludeLabel = ttk.Label(self.topologyFrame, text="Config 포함 : ")
		self.configIncludeLabel.grid(row=6, column=1, padx=(8, 8), pady=(8, 8))
		self.configInclude = IntVar()
		self.configInclude.set(1)
		self.configIncludeCheckBox = ttk.Checkbutton(self.topologyFrame, width=self.defaultComboBoxW, variable=self.configInclude)
		self.configIncludeCheckBox.grid(row=6, column=2, padx=(8, 8), pady=(8, 8))
  
  
		self.switchIconLabel = ttk.Label(self.topologyFrame, text="Switch Icon : ")
		self.switchIconLabel.grid(row=7, column=1)
		self.switchIconComboBox = ttk.Combobox(self.topologyFrame, width=self.defaultComboBoxW, values=self.switchIcon, state="readonly")
		self.switchIconComboBox.current(0)
		self.switchIconComboBox.grid(row=7, column=2, padx=(8, 8), pady=(8, 8))
  
		self.topologyButton = tk.Button(self.topologyFrame, text="Lab Topology 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createTopology()))
		self.topologyButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### 백업
		self.initFrame = ttk.Frame(self.root)
		self.notebook.add(self.initFrame, text=" 백업 및 초기화 ")
  
		self.backupLabel = ttk.Label(self.initFrame, text="백업메모 : ")
		self.backupLabel.grid(row=1, column=1)
		self.backup = StringVar()
		self.backupTextBox = ttk.Entry(self.initFrame, width=self.defaultbackupTextBoxW, textvariable=self.backup)
		self.backupTextBox.grid(row=1, column=2)
  
		self.buttonBackUp = tk.Button(self.initFrame, text="Config 백업", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.backupConfigCall()))
		self.buttonBackUp.grid(row=1, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.initFrame, text="Config 백업 List", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.getBackupConfigList()))
		self.buttonList.grid(row=2, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="Config 백업 선택 배포", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.selectedConfigCall()))
		self.buttonSelectedDeploy.grid(row=3, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="-------------------", width=self.defaultButtonW)
		self.buttonSelectedDeploy.grid(row=4, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		
		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="※※※ 초기화 ※※※", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.cleanConfigCall()))
		self.buttonSelectedDeploy.grid(row=5, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### 핑테스트
		self.pingTestFrame = ttk.Frame(self.root)
		self.notebook.add(self.pingTestFrame, text=" P2P PING TEST ")
		
		self.pingTestButton = tk.Button(self.pingTestFrame, text="P2P PING TEST", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.pingTest()))
		self.pingTestButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		##### 간편 topology 생성
		self.simpleTopologyFrame = ttk.Frame(self.root)
		self.notebook.add(self.simpleTopologyFrame, text=" 간편 LAB-TOPOLOGY ")
  
  
		self.simpleTopologyLabel = ttk.Label(self.simpleTopologyFrame, text="이름(영문) : ")
		self.simpleTopologyLabel.grid(row=0, column=1, padx=(8, 8), pady=(8, 8))
		self.simpleTopology = StringVar()
		self.simpleTopologyTextBox = ttk.Entry(self.simpleTopologyFrame, width=self.defaultTextBoxW, textvariable=self.simpleTopology)
		self.simpleTopologyTextBox.grid(row=0, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleVersionLabel = ttk.Label(self.simpleTopologyFrame, text="버전 : ")
		self.simpleVersionLabel.grid(row=1, column=1, padx=(8, 8), pady=(8, 8))
		self.simpleVersion = StringVar()
		self.simpleVersion.set("veos-4.27.4.1M-noztp")
		self.simpleVersionTextBox = ttk.Entry(self.simpleTopologyFrame, width=self.defaultTextBoxW, textvariable=self.simpleVersion)
		self.simpleVersionTextBox.grid(row=1, column=2, padx=(8, 8), pady=(8, 8))

		self.simpleSpinePrefixLabel = ttk.Label(self.simpleTopologyFrame, text="Spine Name : ")
		self.simpleSpinePrefixLabel.grid(row=2, column=1, padx=(8, 8), pady=(8, 8))
		self.simpleSpinePrefix = StringVar()
		self.simpleSpinePrefix.set("Spine-")
		self.simpleSpinePrefixTextBox = ttk.Entry(self.simpleTopologyFrame, width=self.defaultTextBoxW, textvariable=self.simpleSpinePrefix)
		self.simpleSpinePrefixTextBox.grid(row=2, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleSpineLabel = ttk.Label(self.simpleTopologyFrame, text="Spine 개수 : ")
		self.simpleSpineLabel.grid(row=3, column=1, padx=(8, 8), pady=(8, 8))
		# self.simpleSpine = StringVar()
		self.simpleSpineComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.spineSize, state="readonly")
		self.simpleSpineComboBox.current(0)
		self.simpleSpineComboBox.grid(row=3, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleLeafPrefixLabel = ttk.Label(self.simpleTopologyFrame, text="Leaf Name : ")
		self.simpleLeafPrefixLabel.grid(row=4, column=1, padx=(8, 8), pady=(8, 8))
		self.simpleLeafPrefix = StringVar()
		self.simpleLeafPrefix.set("Leaf-")
		self.simpleLeafPrefixTextBox = ttk.Entry(self.simpleTopologyFrame, width=self.defaultTextBoxW, textvariable=self.simpleLeafPrefix)
		self.simpleLeafPrefixTextBox.grid(row=4, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleLeafLabel = ttk.Label(self.simpleTopologyFrame, text="Leaf 개수 : ")
		self.simpleLeafLabel.grid(row=5, column=1, padx=(8, 8), pady=(8, 8))
		# self.simpleLeaf = StringVar()
		self.simpleLeafComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.leafSize, state="readonly")
		self.simpleLeafComboBox.current(6)
		self.simpleLeafComboBox.grid(row=5, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleEthernetLabel = ttk.Label(self.simpleTopologyFrame, text="Ethernet 개수 : ")
		self.simpleEthernetLabel.grid(row=6, column=1)
		self.simpleEthernetComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.ethernetSize, state="readonly")
		self.simpleEthernetComboBox.current(0)
		self.simpleEthernetComboBox.grid(row=6, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleCpuLabel = ttk.Label(self.simpleTopologyFrame, text="CPU 개수 : ")
		self.simpleCpuLabel.grid(row=7, column=1)
		self.simpleCpuComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.cpuSize, state="readonly")
		self.simpleCpuComboBox.current(1)
		self.simpleCpuComboBox.grid(row=7, column=2, padx=(8, 8), pady=(8, 8))
  
		
		self.simpleRamLabel = ttk.Label(self.simpleTopologyFrame, text="RAM 크기 : ")
		self.simpleRamLabel.grid(row=8, column=1)
		self.simpleRamComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.ramSize, state="readonly")
		self.simpleRamComboBox.current(3)
		self.simpleRamComboBox.grid(row=8, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleCloudLabel = ttk.Label(self.simpleTopologyFrame, text="Cloud1 네트워크 생성 : ")
		self.simpleCloudLabel.grid(row=9, column=1, padx=(8, 8), pady=(8, 8))
		self.simpleCloud = IntVar()
		self.simpleCloudCheckBox = ttk.Checkbutton(self.simpleTopologyFrame, width=self.defaultComboBoxW, variable=self.simpleCloud)
		self.simpleCloudCheckBox.grid(row=9, column=2, padx=(8, 8), pady=(8, 8))
  
		self.simpleSwitchIconLabel = ttk.Label(self.simpleTopologyFrame, text="Switch Icon : ")
		self.simpleSwitchIconLabel.grid(row=10, column=1)
		self.simpleSwitchIconComboBox = ttk.Combobox(self.simpleTopologyFrame, width=self.defaultComboBoxW, values=self.switchIcon, state="readonly")
		self.simpleSwitchIconComboBox.current(0)
		self.simpleSwitchIconComboBox.grid(row=10, column=2, padx=(8, 8), pady=(8, 8))


  
		self.simpleTopologyButton = tk.Button(self.simpleTopologyFrame, text="Lab Topology 생성", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.createSimpleTopology()))
		self.simpleTopologyButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### EVE-NG, PNET os 이미지 설정
		self.labOsFrame = ttk.Frame(self.root)
		self.notebook.add(self.labOsFrame, text=" LAB OS ")
		
		self.labOsVendorLabel = ttk.Label(self.labOsFrame, text="제조사 : ")
		self.labOsVendorLabel.grid(row=0, column=1)
		self.labOsVender = ttk.Combobox(self.labOsFrame, width=self.defaultComboBoxW, values=self.vendors, state="readonly")
		self.labOsVender.current(0)
		self.labOsVender.grid(row=0, column=2, padx=(8, 8), pady=(8, 8))
  
		self.labServerIpLabel = ttk.Label(self.labOsFrame, text="Lab Server IP : ")
		self.labServerIpLabel.grid(row=1, column=1, padx=(8, 8), pady=(8, 8))
		self.labServerIp = StringVar()
		self.labServerIpTextBox = ttk.Entry(self.labOsFrame, width=self.defaultTextBoxW, textvariable=self.labServerIp)
		self.labServerIpTextBox.grid(row=1, column=2, padx=(8, 8), pady=(8, 8))
  
		self.labServerIdLabel = ttk.Label(self.labOsFrame, text="Lab Server ID: ")
		self.labServerIdLabel.grid(row=2, column=1, padx=(8, 8), pady=(8, 8))
		self.labServerId = StringVar()
		self.labServerIdTextBox = ttk.Entry(self.labOsFrame, width=self.defaultTextBoxW, textvariable=self.labServerId)
		self.labServerIdTextBox.grid(row=2, column=2, padx=(8, 8), pady=(8, 8))
  
		self.labServerPwLabel = ttk.Label(self.labOsFrame, text="Lab Server PW: ")
		self.labServerPwLabel.grid(row=3, column=1, padx=(8, 8), pady=(8, 8))
		self.labServerPw = StringVar()
		self.labServerPwTextBox = ttk.Entry(self.labOsFrame, width=self.defaultTextBoxW, textvariable=self.labServerPw)
		self.labServerPwTextBox.grid(row=3, column=2, padx=(8, 8), pady=(8, 8))
		self.labServerPwTextBox.config(show="*")
  
		self.labServerPortLabel = ttk.Label(self.labOsFrame, text="Lab Server SSH Port: ")
		self.labServerPortLabel.grid(row=4, column=1, padx=(8, 8), pady=(8, 8))
		self.labServerPort = StringVar()
		self.labServerPort.set(22)
		self.labServerPortTextBox = ttk.Entry(self.labOsFrame, width=self.defaultTextBoxW, textvariable=self.labServerPort)
		self.labServerPortTextBox.grid(row=4, column=2, padx=(8, 8), pady=(8, 8))
  
		self.labOsListButton = tk.Button(self.labOsFrame, text="OS 목록", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.getLabOsList()))
		self.labOsListButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.labOsListButton = tk.Button(self.labOsFrame, text="Lab OS 추가", width=self.defaultButtonW, command=lambda: self.loop.create_task(self.appendLabOS()))
		self.labOsListButton.grid(row=1, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.root.protocol("WM_DELETE_WINDOW", self.windowButtonClose)

		self.asyncio.get_event_loop_policy().get_event_loop().run_until_complete(self.mainLoop())
	
	async def mainLoop(self):
		while True:
			try:
				self.root.winfo_exists()  # Will throw TclError if the main window is destroyed
				self.root.update()
			except TclError:
				break

			await self.asyncio.sleep(0.01)

	def resource_path(self, relative_path):
		try:
			# PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수
			base_path = sys._MEIPASS
		except Exception:
			base_path = os.path.abspath(".")
		return os.path.join(base_path, relative_path)
  
	def ymlInit(self):
		self.processAuto.ymlInit()
  
	def norInit(self):
		self.nr = self.processAuto.norInit()
  
	def windowButtonClose(self):
		try:
			self.loop.complete()
			self.loop.stop()
			self.loop.close()
		except:
			print("close exception!")		
		finally:
			self.root.quit()
			self.root.destroy()
			exit()
   
	  
	async def cleanConfigCall(self):
  
		answer = askyesno(title='confirmation',
                    message='초기화를 진행하시겠습니까?. 스위치의 모든 정보가 초기화 됩니다')
		if not answer:
			pass
		else:
			self.gridReset()
			self.nr = self.processAuto.cleanConfigCall()
		
			for host in self.nr.inventory.hosts:
				if host in self.nr.data.failed_hosts:
					desc = "초기화 실패"
				else:
					desc = "초기화 완료"

				self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
  
	async def getBackupConfigList(self):
		self.gridReset()
  
		sheet = self.processAuto.getBackupConfigList()
  
		if sheet.iter_rows():
			self.lastConfigGen = "backup"
			for row in sheet.iter_rows():
				self.treeview.insert('', END, text='', values=(row[1].value, row[0].value, row[2].value))
		else:
			messagebox.showwarning(title="info", message="백업내역이 없습니다.")

	async def selectedConfigCall(self):

		if self.lastConfigGen != "backup":
			messagebox.showwarning(title="warning", message="Config 백업 List 호출후 사용하실 수 있습니다.")
			return False
   
		iid = self.treeview.focus()
		if iid:
			item = self.treeview.item(iid).get("values")[0]
			desc = self.treeview.item(iid).get("values")[1]
		else:
			messagebox.showwarning(title="warning", message="선택된 영역이 없습니다.")
			return FALSE
 
		directory = self.path + f"inventory/config_backup/{item}"
		# print(f"config_backup/{item}")
		if not os.path.exists(directory):
			messagebox.showwarning(title="warning", message="백업내역이 없습니다. 확인후 다시 시도하세요.")
			return FALSE
 
		
		self.gridReset()
		self.nr = self.processAuto.selectedConfigCall(item)
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				descDetail = desc + " config 배포 실패"
			else:
				descDetail = desc + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, descDetail, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.nr = self.processAuto.norInit()
		
		# print("done!!")
	
	async def sendConfigCall(self):
		"""
		config 배포 호출
		"""
		
		if not self.processAuto.sendConfigCheck():
			messagebox.showwarning(title="warning", message="config 생성 후 사용하실 수 있습니다.")
			return False
 
		self.gridReset()
		self.nr = self.processAuto.sendConfigCall()
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = self.processAuto.getLastConfigGen() + " config 배포 실패"
			else:
				desc = self.processAuto.getLastConfigGen() + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.nr = self.processAuto.norInit()
  
	def gridReset(self):
		"""
		Grid 내용 초기화
		"""
		x = self.treeview.get_children()
		for item in x:
			self.treeview.delete(item)
	
		# self.nr.data.reset_failed_hosts()
   
	async def show(self):
		
		while True:
			
			self.root.update()
					
			await asyncio.sleep(.01)


	async def backupConfigCall(self):
		self.gridReset()
		
		memo = self.backup.get()
		self.backup.set("")
  
		if memo == "":
			memo = "config 백업"
   
   
		nowDate = datetime.now().strftime('%y-%m-%d %H:%M:%S')
		self.nr = self.processAuto.backupConfigCall(memo)
	

		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "Config 백업 실패"
			else:
				desc = "Config 백업 완료"

			self.treeview.insert('', END, text='', values=(host, desc, nowDate))

	async def createTopology(self):
   
		cpu = self.cpuComboBox.get()
		version = self.version.get()
		ram = self.ramComboBox.get()
		topology = self.topology.get()
		ethernetCount = self.ethernetComboBox.get()
		net = self.cloud.get()
		icon = self.switchIconComboBox.get()
		configInclude = self.configInclude.get()
  
		if eq(net, 1):
			net = True
		else:
			net = False
   
		if eq(configInclude, 1):
			config = True
		else:
			config = False

		# print("----------------------")
		# print(cpu, version, ram, topology, ethernetCount)
  
		if eq("", topology):
			messagebox.showwarning(title="warning", message="토폴로지이름을 입력해 주세요.")
			self.topologyTextBox.focus()
			return False
 
		if eq("", version):
			messagebox.showwarning(title="warning", message="버전을 입력해 주세요.")
			self.versionTextBox.focus()
			return False

		if config:
			session = self.processAuto.getLastConfigGen()
			print("session = ", session)
			if not session in self.sessions:
				messagebox.showwarning(title="warning", message="Config 생성후 생성해주세요")
				return False

		# if eq("", ethernetCount):
		# 	messagebox.showwarning(title="warning", message="ethernet 갯수를 입력해 주세요.")
		# 	self.ethernetComboBox.focus()
		# 	return False
    
		# if eq("", cpu):
		# 	messagebox.showwarning(title="warning", message="cpu 갯수를 입력해 주세요.")
		# 	self.cpuComboBox.focus()
		# 	return False
    
		# if eq("", ram):
		# 	messagebox.showwarning(title="warning", message="ram 사이즈를 입력해 주세요.")
		# 	self.ramComboBox.focus()
		# 	return False
   
		self.gridReset()
  
		self.processAuto.createTopology(topology=topology, cpu=cpu, ram=ram, version=version, ethernetCount=ethernetCount, net=net, icon=icon, configInclude=configInclude)
  
		self.treeview.insert('', END, text='', values=("topology", topology + "이(가) 생성되었습니다.", datetime.now().strftime('%y-%m-%d %H:%M:%S')))
  
	async def createSimpleTopology(self):
   
		cpu = self.simpleCpuComboBox.get()
		version = self.simpleVersion.get()
		ram = self.simpleRamComboBox.get()
		topology = self.simpleTopology.get()
		ethernetCount = int(self.simpleEthernetComboBox.get())
		spineCount = int(self.simpleSpineComboBox.get())
		leafCount = int(self.simpleLeafComboBox.get())
		spinePrefix = self.simpleSpinePrefix.get()
		leafPrefix = self.simpleLeafPrefix.get()
		net = self.simpleCloud.get()
		icon = self.simpleSwitchIconComboBox.get()
		if eq(net, 1):
			net = True
		else:
			net = False
   
		if eq("", topology):
			messagebox.showwarning(title="warning", message="토폴로지이름을 입력해 주세요.")
			self.simpleTopologyTextBox.focus()
			return False
 
		if eq("", version):
			messagebox.showwarning(title="warning", message="버전을 입력해 주세요.")
			self.simpleVersionTextBox.focus()
			return False
 
		if eq("", spinePrefix):
			messagebox.showwarning(title="warning", message="spine Name을 입력해 주세요.")
			self.simpleSpinePrefixTextBox.focus()
			return False

		if eq("", leafPrefix):
			messagebox.showwarning(title="warning", message="leaf Name을 입력해 주세요.")
			self.simpleLeafPrefixTextBox.focus()
			return False
 
		if eq("", ethernetCount):
			messagebox.showwarning(title="warning", message="ethernet 갯수를 입력해 주세요.")
			self.simpleEthernetComboBox.focus()
			return False
    
    
		if (ethernetCount < leafCount):
			messagebox.showwarning(title="warning", message="ethernet 갯수가 leaf 갯수보다 작습니다.")
			return False

		if eq("", cpu):
			messagebox.showwarning(title="warning", message="cpu 갯수를 입력해 주세요.")
			self.simpleCpuComboBox.focus()
			return False
    
		if eq("", ram):
			messagebox.showwarning(title="warning", message="ram 사이즈를 입력해 주세요.")
			self.simpleRamComboBox.focus()
			return False
   
		self.gridReset()
  
		self.processAuto.createSimpleTopology(topology=topology, spinePrefix=spinePrefix, leafPrefix=leafPrefix, spinesCount=spineCount, leafsCount=leafCount, cpu=cpu, ram=ram, version=version, ethernetCount=ethernetCount, net=net, icon=icon)
		self.treeview.insert('', END, text='', values=("간편 topology", topology + "이(가) 생성되었습니다.", datetime.now().strftime('%y-%m-%d %H:%M:%S')))
		
    

	async def createConfig(self, session):
		self.nr = self.processAuto.createConfig(session)
   
		self.gridReset()
   
		for host in self.nr.inventory.hosts:
    
			self.treeview.insert('', END, text='', values=(host, session + " config 생성 완료", datetime.now().strftime('%y-%m-%d %H:%M:%S')))

	async def statusCheckCall(self):
   
		self.gridReset()
		statusResult = self.processAuto.statusCheckCall()

		for host in statusResult:
			try:
				result = dict(statusResult[host][0].result)
				model = result["model"].replace("DCS-", "")
				timezone = result["timezone"]
				date = result["date"]
				version = result["version"]

				desc = f"{model}, {version}, {timezone}, {date}"
			except ValueError:
				desc = "connect... fail"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))

	async def pingTest(self):
		self.gridReset()
  
		pingTestResult = self.processAuto.pingTestCall()
  
		for host in pingTestResult:
			result = dict(pingTestResult[host][0].result)
			success = result["success"]
			fail = result["fail"]
			down = result["down"]
			self.treeview.insert('', END, text='', values=(host, f'ping test !! pass: {success}, fail: {fail}, down: {down}', datetime.now().strftime('%y-%m-%d %H:%M:%S')))

	async def getLabOsList(self):
		vendor = self.labOsVender.get()
		osList, result = self.processLab.getLabOsList(vendor)
		self.gridReset()
  
		if eq(result, 0):
			messagebox.showwarning(title="warning", message="파일서버가 접속되지 않습니다. 확인후 다시 시도해 주세요.")
			return False

		if not osList:
			messagebox.showwarning(title="warning", message="os 파일이 없습니다. 확인후 다시 시도해 주세요.")
			return False
  
		for os in osList:
			os = str(os).replace(".qcow2", "")
			self.treeview.insert('', END, text='', values=(vendor, os, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
	async def appendLabOS(self):
		ip = self.labServerIp.get()
		id = self.labServerId.get()
		pw = self.labServerPw.get()
		port = self.labServerPort.get()
		iid = self.treeview.focus()
  
		if iid:
			vendor = self.treeview.item(iid).get("values")[0]
			osVersion = self.treeview.item(iid).get("values")[1]
		else:
			messagebox.showwarning(title="warning", message="선택된 OS 버전이 없습니다.")
			return False
 
		if eq(ip, ""):
			messagebox.showwarning(title="warning", message="Lab Server IP를 입력해 주세요.")
			self.labServerIpTextBox.focus()
			return False

 
		if eq(id, ""):
			messagebox.showwarning(title="warning", message="Lab Server ID를 입력해 주세요.")
			self.labServerIdTextBox.focus()
			return False
 
		if eq(pw, ""):
			messagebox.showwarning(title="warning", message="Lab Server PW를 입력해 주세요.")
			self.labServerPwTextBox.focus()
			return False
 
		if eq(port, ""):
			messagebox.showwarning(title="warning", message="Lab Server SSH Port를 입력해 주세요.")
			self.labServerPortTextBox.focus()
			return False

		print(vendor, osVersion)
		
		result, msg = self.processLab.labOSAppend(vendor=vendor, osVersion=osVersion, host=ip, user=id, pw=pw, port=port)

		if eq(result, -1):
			messagebox.showwarning(title="warning", message="Lab Server 접속에 실패하였습니다. IP/ID/PW 확인해주세요.")
			return False
		elif eq(result, -2):
			messagebox.showwarning(title="warning", message=msg)
			return False
 
		self.gridReset()
		self.treeview.insert('', END, text='', values=("LAB OS", f'{vendor} OS {osVersion} 추가 완료', datetime.now().strftime('%y-%m-%d %H:%M:%S')))