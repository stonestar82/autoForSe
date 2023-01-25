import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import *
import yaml, json
import asyncio, os
from nornir import InitNornir
from nornir.plugins import *
from nornir.plugins.inventory import *
from nornir.plugins.runners import *
from nornir.plugins.functions import *
from nornir.plugins.processors import *
from nornir.plugins.tasks import *
from nornir.plugins.connections import *
from nornir_netmiko.tasks import netmiko_send_config, netmiko_send_command
from nornir.core.task import Task, Result
from jinja2 import Template
from datetime import datetime
from openpyxl import load_workbook
from generators.BlankNone import *
from generators.generateInventory import generateInventory 
from operator import eq
import zipfile
from src.ProcessImpl import ProcessImpl
from tkinter.messagebox import askyesno
import platform

class UI2():
	def __init__(self, asyncio=asyncio, process=ProcessImpl):
		self.asyncio = asyncio
		self.process = process
		# self.loop = self.asyncio.get_event_loop()
		# if platform.system()=='Windows':
		# 	asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
		self.root = tk.Tk()
		self.root.title("GlobalTelecom & i-Cloud")
		self.root.geometry("640x500+600+100") ## w, h, x, y
		self.root.resizable(False, False)
  
		self.db = "./db/db.xlsx"
		self.inventory = "./inventory.xlsx"
		self.cpuSize = [1,2]
		self.ethernetSize = []
		for i in range(9, 97):
			self.ethernetSize.append(i)
   
		self.spineSize = [2,4]
		self.leafSize = []
		for i in range(2, 60):
			self.leafSize.append(i)
   
		self.ramSize = [1024, 2048, 3072, 4096]
		self.lastConfigGen = "" ## 마지막으로 실행한 config 생성 session
		self.fullSession = {"fullv4":["init", "base", "loop0", "p2pip", "bgpv4", "etcport", "vxlan"] , "fullv6": ["init", "base", "loop0", "p2pipv6", "bgpv6", "etcport", "vxlan"]}
		self.sessions = list(set(self.fullSession["fullv4"] + self.fullSession["fullv6"]))
  
		self.ymlInit()
		self.norInit()
  
  
		##### grid S #####
		self.frame_top = ttk.Frame(self.root)
		self.frame_top.pack(side="top")


		self.treeview=ttk.Treeview(self.frame_top, columns=["one", "two", "three"])
		self.treeview.tag_configure("green",foreground='green')
		self.treeview.tag_configure("red",foreground='red')

		self.treeview.column("#1", width=180, anchor="center")
		self.treeview.heading("one", text="name", anchor="center")

		self.treeview.column("#2", width=330, anchor="w")
		self.treeview.heading("two", text="description", anchor="center")
  
		self.treeview.column("#3", width=120, anchor="w")
		self.treeview.heading("three", text="datetime", anchor="center")


		self.treeview["show"] = "headings"
		self.treeview.pack()
		##### grid E #####

		##### tab 구현 S #####
		self.notebook = ttk.Notebook(self.root, width=620, height=300, )
		self.notebook.pack()
  
	
		##### ipv4
		self.ipv4Frame = ttk.Frame(self.root)
		self.notebook.add(self.ipv4Frame, text="IPV4")

		self.buttonStatus = tk.Button(self.ipv4Frame, text="init Config 생성", width=18, command=lambda: self.createConfig("init"))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.ipv4Frame, text="base Config 생성", width=18, command=lambda: self.createConfig("base"))
		self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.ipv4Frame, text="loop0 Config 생성", width=18, command=lambda: self.createConfig("loop0"))
		self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.ipv4Frame, text="etcport Config 생성", width=18, command=lambda: self.createConfig("etcport"))
		self.buttonList.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.ipv4Frame, text="vxlan Config 생성", width=18, command=lambda: self.createConfig("vxlan"))
		self.buttonList.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonStatus = tk.Button(self.ipv4Frame, text="p2p IPv4 생성", width=18, command=lambda: self.createConfig("p2pip"))
		self.buttonStatus.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.ipv4Frame, text="BGP IPv4 생성", width=18, command=lambda: self.createConfig("bgpv4"))
		self.buttonBackUp.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
	
		self.buttonStatus = tk.Button(self.ipv4Frame, text="full IPv4 Config 생성", width=18, command=lambda: self.createConfig("fullv4"))
		self.buttonStatus.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.ipv4Frame, text="Config 배포", width=18, command=lambda: self.sendConfigCall())
		self.buttonSelectedDeploy.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### ipv6
		self.ipv6Frame = ttk.Frame(self.root)
		self.notebook.add(self.ipv6Frame, text="IPV6")

		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="init Config 생성", width=18, command=lambda: self.createConfig("init"))
		self.buttonStatusIpv6.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUpIpv6 = tk.Button(self.ipv6Frame, text="base Config 생성", width=18, command=lambda: self.createConfig("base"))
		self.buttonBackUpIpv6.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="loop0 Config 생성", width=18, command=lambda: self.createConfig("loop0"))
		self.buttonListIpv6.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="etcport Config 생성", width=18, command=lambda: self.createConfig("etcport"))
		self.buttonListIpv6.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonListIpv6 = tk.Button(self.ipv6Frame, text="vxlan Config 생성", width=18, command=lambda: self.createConfig("vxlan"))
		self.buttonListIpv6.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="p2p IPv6 생성", width=18, command=lambda: self.createConfig("p2pipv6"))
		self.buttonStatusIpv6.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUpIpv6 = tk.Button(self.ipv6Frame, text="BGP IPv6 생성", width=18, command=lambda: self.createConfig("bgpv6"))
		self.buttonBackUpIpv6.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonStatusIpv6 = tk.Button(self.ipv6Frame, text="full IPv6 Config 생성", width=18, command=lambda: self.createConfig("fullv6"))
		self.buttonStatusIpv6.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		self.buttonSelectedDeployIpv6 = tk.Button(self.ipv6Frame, text="Config 배포", width=18, command=lambda: self.loop.create_task(self.sendConfigCall()))
		self.buttonSelectedDeployIpv6.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		##### topology
		self.topologyFrame = ttk.Frame(self.root)
		self.notebook.add(self.topologyFrame, text="topology")
  
  
		self.topologyLabel = ttk.Label(self.topologyFrame, text="이름(영문) : ")
		self.topologyLabel.grid(row=0, column=1)
		self.topology = StringVar()
		self.topologyTextBox = ttk.Entry(self.topologyFrame, width=40, textvariable=self.topology)
		self.topologyTextBox.grid(row=0, column=2)
  
		self.versionLabel = ttk.Label(self.topologyFrame, text="버전 : ")
		self.versionLabel.grid(row=1, column=1)
		self.version = StringVar()
		self.versionTextBox = ttk.Entry(self.topologyFrame, width=40, textvariable=self.version)
		self.versionTextBox.grid(row=1, column=2)

		self.ethernetLabel = ttk.Label(self.topologyFrame, text="ethernet 개수 : ")
		self.ethernetLabel.grid(row=2, column=1)
		# self.ethernet = StringVar()
		self.ethernetComboBox = ttk.Combobox(self.topologyFrame, width=38, values=self.ethernetSize, state="readonly")
		self.ethernetComboBox.current(0)
		self.ethernetComboBox.grid(row=2, column=2)
  
		self.cpuLabel = ttk.Label(self.topologyFrame, text="cpu 개수 : ")
		self.cpuLabel.grid(row=3, column=1)
		# self.cpu = StringVar()
		self.cpuComboBox = ttk.Combobox(self.topologyFrame, width=38, values=self.cpuSize, state="readonly")
		self.cpuComboBox.current(1)
		self.cpuComboBox.grid(row=3, column=2)
  
		
		self.ramLabel = ttk.Label(self.topologyFrame, text="ram 크기 : ")
		self.ramLabel.grid(row=4, column=1)
		# self.ram = StringVar()
		self.ramComboBox = ttk.Combobox(self.topologyFrame, width=38, values=self.ramSize, state="readonly")
		self.ramComboBox.current(3)
		self.ramComboBox.grid(row=4, column=2)
  
		self.topologyButton = tk.Button(self.topologyFrame, text="토폴로지 생성", width=18, command=lambda: self.createTopology())
		self.topologyButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
		##### 백업
		self.initFrame = ttk.Frame(self.root)
		self.notebook.add(self.initFrame, text="백업 및 초기화")
  
		self.backupLabel = ttk.Label(self.initFrame, text="백업메모 : ")
		self.backupLabel.grid(row=1, column=1)
		self.backup = StringVar()
		self.backupTextBox = ttk.Entry(self.initFrame, width=50, textvariable=self.backup)
		self.backupTextBox.grid(row=1, column=2)
  
		self.buttonBackUp = tk.Button(self.initFrame, text="Config 백업", width=18, command=lambda: self.backupConfigCall())
		self.buttonBackUp.grid(row=1, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.initFrame, text="Config 백업 List", width=18, command=lambda: self.getBackupConfigList())
		self.buttonList.grid(row=2, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="Config 백업 선택 배포", width=18, command=lambda: self.selectedConfigCall())
		self.buttonSelectedDeploy.grid(row=3, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="-------------------", width=18)
		self.buttonSelectedDeploy.grid(row=4, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		
		self.buttonSelectedDeploy = tk.Button(self.initFrame, text="※※※ 초기화 ※※※", width=18, command=lambda: self.cleanConfigCall())
		self.buttonSelectedDeploy.grid(row=5, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
  
  
		##### 간편 topology 생성
		self.simpleTopologyFrame = ttk.Frame(self.root)
		self.notebook.add(self.simpleTopologyFrame, text="간편 topology")
  
  
		self.simpleTopologyLabel = ttk.Label(self.simpleTopologyFrame, text="이름(영문) : ")
		self.simpleTopologyLabel.grid(row=0, column=1)
		self.simpleTopology = StringVar()
		self.simpleTopologyTextBox = ttk.Entry(self.simpleTopologyFrame, width=40, textvariable=self.simpleTopology)
		self.simpleTopologyTextBox.grid(row=0, column=2)
  
		self.simpleVersionLabel = ttk.Label(self.simpleTopologyFrame, text="버전 : ")
		self.simpleVersionLabel.grid(row=1, column=1)
		self.simpleVersion = StringVar()
		self.simpleVersionTextBox = ttk.Entry(self.simpleTopologyFrame, width=40, textvariable=self.simpleVersion)
		self.simpleVersionTextBox.grid(row=1, column=2)

		self.simpleSpinePrefixLabel = ttk.Label(self.simpleTopologyFrame, text="spine prefix : ")
		self.simpleSpinePrefixLabel.grid(row=2, column=1)
		self.simpleSpinePrefix = StringVar()
		self.simpleSpinePrefixTextBox = ttk.Entry(self.simpleTopologyFrame, width=40, textvariable=self.simpleSpinePrefix)
		self.simpleSpinePrefixTextBox.grid(row=2, column=2)
  
		self.simpleSpineLabel = ttk.Label(self.simpleTopologyFrame, text="spine 개수 : ")
		self.simpleSpineLabel.grid(row=3, column=1)
		# self.simpleSpine = StringVar()
		self.simpleSpineComboBox = ttk.Combobox(self.simpleTopologyFrame, width=38, values=self.spineSize, state="readonly")
		self.simpleSpineComboBox.current(0)
		self.simpleSpineComboBox.grid(row=3, column=2)
  
		self.simpleLeafPrefixLabel = ttk.Label(self.simpleTopologyFrame, text="leaf prefix : ")
		self.simpleLeafPrefixLabel.grid(row=4, column=1)
		self.simpleLeafPrefix = StringVar()
		self.simpleLeafPrefixTextBox = ttk.Entry(self.simpleTopologyFrame, width=40, textvariable=self.simpleLeafPrefix)
		self.simpleLeafPrefixTextBox.grid(row=4, column=2)
  
		self.simpleLeafLabel = ttk.Label(self.simpleTopologyFrame, text="leaf 개수 : ")
		self.simpleLeafLabel.grid(row=5, column=1)
		# self.simpleLeaf = StringVar()
		self.simpleLeafComboBox = ttk.Combobox(self.simpleTopologyFrame, width=38, values=self.leafSize, state="readonly")
		self.simpleLeafComboBox.current(6)
		self.simpleLeafComboBox.grid(row=5, column=2)
  
		self.simpleEthernetLabel = ttk.Label(self.simpleTopologyFrame, text="ethernet 개수 : ")
		self.simpleEthernetLabel.grid(row=6, column=1)
		# self.simpleEthernet = StringVar()
		self.simpleEthernetComboBox = ttk.Combobox(self.simpleTopologyFrame, width=38, values=self.ethernetSize, state="readonly")
		self.simpleEthernetComboBox.current(0)
		self.simpleEthernetComboBox.grid(row=6, column=2)
  
		self.simpleCpuLabel = ttk.Label(self.simpleTopologyFrame, text="cpu 개수 : ")
		self.simpleCpuLabel.grid(row=7, column=1)
		# self.simpleCpu = StringVar()
		self.simpleCpuComboBox = ttk.Combobox(self.simpleTopologyFrame, width=38, values=self.cpuSize, state="readonly")
		self.simpleCpuComboBox.current(1)
		self.simpleCpuComboBox.grid(row=7, column=2)
  
		
		self.simpleRamLabel = ttk.Label(self.simpleTopologyFrame, text="ram 크기 : ")
		self.simpleRamLabel.grid(row=8, column=1)
		# self.simpleRam = StringVar()
		self.simpleRamComboBox = ttk.Combobox(self.simpleTopologyFrame, width=38, values=self.ramSize, state="readonly")
		self.simpleRamComboBox.current(3)
		self.simpleRamComboBox.grid(row=8, column=2)
  
		self.simpleTopologyButton = tk.Button(self.simpleTopologyFrame, text="간편 토폴로지 생성", width=18, command=lambda: self.createSimpleTopology())
		self.simpleTopologyButton.grid(row=0, column=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.root.mainloop()
		##### tab 구현 E #####
  
		##### 메모 S #####
		# self.frame_middle = ttk.Frame(self.root)
		# self.frame_middle.pack(side="top")
  
		# self.iddLabel = ttk.Label(self.frame_middle, text="메모")
		# self.iddLabel.grid(column=0, row=1)
		# self.memo = StringVar()
		# self.memoTextbox = ttk.Entry(self.frame_middle, width=50, textvariable=self.memo)
		# self.memoTextbox.grid(column=1 , row=1)
		# self.memoTextbox.focus()  
		##### 메모 E #####
	  
		##### 백업 및 grid 선택 배포 S #####
		# self.frameLeft = ttk.Frame(self.root)
		# self.frameLeft.pack(side="left")
  
		# self.buttonStatus = tk.Button(self.frameLeft, text="상태체크", width=18, command=lambda: self.statusCheckCall()))
		# self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameLeft, text="Config 백업", width=18, command=lambda: self.backupConfigCall()))
		# self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		# self.buttonList = tk.Button(self.frameLeft, text="Config 백업 List", width=18, command=lambda: self.getBackupConfigList()))
		# self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		# self.buttonSelectedDeploy = tk.Button(self.frameLeft, text="Config 백업 선택 배포", width=18, command=lambda: self.selectedConfigCall()))
		# self.buttonSelectedDeploy.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### 백업 및 grid 선택 배포 E #####
  
  
		##### session 별 config 생성 및 배포 S #####
		# self.frameMiddle = ttk.Frame(self.root)
		# self.frameMiddle.pack(side="left")
  
		# self.buttonStatus = tk.Button(self.frameMiddle, text="init Config 생성", width=18, command=lambda: self.createConfig("init"))
		# self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameMiddle, text="base Config 생성", width=18, command=lambda: self.createConfig("base"))
		# self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		# self.buttonList = tk.Button(self.frameMiddle, text="loop0 Config 생성", width=18, command=lambda: self.createConfig("loop0"))
		# self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonList = tk.Button(self.frameMiddle, text="etcport Config 생성", width=18, command=lambda: self.createConfig("etcport"))
		# self.buttonList.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonList = tk.Button(self.frameMiddle, text="vxlan Config 생성", width=18, command=lambda: self.createConfig("vxlan"))
		# self.buttonList.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonStatus = tk.Button(self.frameMiddle, text="p2p IPv4 생성", width=18, command=lambda: self.createConfig("p2pip"))
		# self.buttonStatus.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameMiddle, text="BGP IPv4 생성", width=18, command=lambda: self.createConfig("bgpv4"))
		# self.buttonBackUp.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		# self.buttonStatus = tk.Button(self.frameMiddle, text="p2p IPv6 생성", width=18, command=lambda: self.createConfig("p2pipv6"))
		# self.buttonStatus.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameMiddle, text="BGP IPv6 생성", width=18, command=lambda: self.createConfig("bgpv6"))
		# self.buttonBackUp.grid(row=3, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		# self.buttonSelectedDeploy = tk.Button(self.frameMiddle, text="Config 배포", width=40, command=lambda: self.sendConfigCall()))
		# self.buttonSelectedDeploy.grid(row=5, columnspan=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### session 별 config 생성 및 배포 E #####

		##### full config 생성 및 배포 S #####
		# self.frameRight = ttk.Frame(self.root)
		# self.frameRight.pack(side="left")
  
		# self.buttonStatus = tk.Button(self.frameRight, text="full IPv4 Config 생성", width=18, command=lambda: self.createConfig("fullv4"))
		# self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameRight, text="full IPv6 Config 생성", width=18, command=lambda: self.createConfig("fullv6"))
		# self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		# self.buttonBackUp = tk.Button(self.frameRight, text="토폴로지 생성", width=18, command=lambda: self.createTopology()))
		# self.buttonBackUp.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		# self.buttonSelectedDeploy = tk.Button(self.frameRight, text="※※※ 초기화 ※※※", width=18, command=lambda: self.cleanConfigCall()))
		# self.buttonSelectedDeploy.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### full config 생성 및 배포 S #####
  
		self.root.protocol("WM_DELETE_WINDOW", self.windowButtonClose)

  
		# self.proccessingState = False
		# self.root.update()
  
	def ymlInit(self):
		self.process.ymlInit()
  
	def norInit(self):
		self.nr = self.process.norInit()
  
	def windowButtonClose(self):
		# print("close!!")
		# exit()
		self.root.quit()
		self.root.destroy()
		# try:
		# 	self.root.destroy()
		# except:
		# 	print("close!")		
		# finally:
		# 	self.root.quit()
		# 	self.root.destroy()
		# 	print("finally close")
   
	def cleanConfig(self, task: Task):
		self.process.cleanConfig()
  
	def cleanConfigCall(self):
  
		answer = askyesno(title='confirmation',
                    message='초기화를 진행하시겠습니까?. 스위치의 모든 정보가 초기화 됩니다')
		if not answer:
			print("no 했다.")
		else:
			self.gridReset()
			print("yes 했다")
			self.nr = self.process.cleanConfig()
		
			for host in self.nr.inventory.hosts:
				if host in self.nr.data.failed_hosts:
					desc = "초기화 실패"
				else:
					desc = "초기화 완료"

				self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
  
	def getBackupConfigList(self):
		self.gridReset()
  
		sheet = self.process.getBackupConfigList()
  
		if sheet.iter_rows():
			self.lastConfigGen = "backup"
			for row in sheet.iter_rows():
				self.treeview.insert('', END, text='', values=(row[1].value, row[0].value, row[2].value))
		else:
			messagebox.showwarning(title="info", message="백업내역이 없습니다.")

	async def selectedConfigCall(self):

		if self.lastConfigGen != "backup":
			messagebox.showwarning(title="warning", message="Config 백업 List 호출후 사용하실 수 있습니다.")
			return False
   
		iid = self.treeview.focus()
		if iid:
			item = self.treeview.item(iid).get("values")[0]
			desc = self.treeview.item(iid).get("values")[1]
		else:
			messagebox.showwarning(title="warning", message="선택된 영역이 없습니다.")
			return FALSE
 
		directory = f"./inventory/config_backup/{item}"
		# print(f"config_backup/{item}")
		if not os.path.exists(directory):
			messagebox.showwarning(title="warning", message="백업내역이 없습니다. 확인후 다시 시도하세요.")
			return FALSE
 
		
		self.gridReset()
		self.nr = self.process.selectedConfigCall(item)
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				descDetail = desc + " config 배포 실패"
			else:
				descDetail = desc + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, descDetail, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.nr = self.process.norInit()
		
		# print("done!!")
	
	def sendConfigCall(self):
		"""
		config 배포 호출
		"""
		print("call!! !!!!")
		if not self.process.sendConfigCheck():
			messagebox.showwarning(title="warning", message="config 생성 후 사용하실 수 있습니다.")
			return False
 
		self.gridReset()
  
		self.nr = self.process.sendConfigCall()
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = self.lastConfigGen + " config 배포 실패"
			else:
				desc = self.lastConfigGen + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.nr = self.process.norInit()
  
	def gridReset(self):
		"""
		Grid 내용 초기화
		"""
		x = self.treeview.get_children()
		for item in x:
			self.treeview.delete(item)
	
		self.nr.data.reset_failed_hosts()
   
	async def show(self):
		
		while True:
			
			self.root.update()
					
			await asyncio.sleep(.01)


	async def backupConfigCall(self):
		self.gridReset()
		  
		nowDate = datetime.now().strftime('%y-%m-%d %H:%M:%S')
		self.nr = self.process.backupConfigCall()
		  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "Config 백업 실패"
			else:
				desc = "Config 백업 완료"

			self.treeview.insert('', END, text='', values=(host, desc, nowDate))

	def createTopology(self):
   
		cpu = self.cpuComboBox.get()
		version = self.version.get()
		ram = self.ramComboBox.get()
		topology = self.topology.get()
		ethernetCount = self.ethernetComboBox.get()
  
		# print("----------------------")
		# print(cpu, version, ram, topology, ethernetCount)
  
		if eq("", topology):
			messagebox.showwarning(title="warning", message="토폴로지이름을 입력해 주세요.")
			self.topologyTextBox.focus()
			return False
 
		if eq("", version):
			messagebox.showwarning(title="warning", message="버전을 입력해 주세요.")
			self.versionTextBox.focus()
			return False

		# if eq("", ethernetCount):
		# 	messagebox.showwarning(title="warning", message="ethernet 갯수를 입력해 주세요.")
		# 	self.ethernetComboBox.focus()
		# 	return False
    
		# if eq("", cpu):
		# 	messagebox.showwarning(title="warning", message="cpu 갯수를 입력해 주세요.")
		# 	self.cpuComboBox.focus()
		# 	return False
    
		# if eq("", ram):
		# 	messagebox.showwarning(title="warning", message="ram 사이즈를 입력해 주세요.")
		# 	self.ramComboBox.focus()
		# 	return False
   
		self.gridReset()
  
		self.process.createTopology(topology=topology, cpu=cpu, ram=ram, version=version, ethernetCount=ethernetCount)
  
		self.treeview.insert('', END, text='', values=("topology", topology + "이(가) 생성되었습니다.", datetime.now().strftime('%y-%m-%d %H:%M:%S')))
  
	def createSimpleTopology(self):
   
		cpu = self.simpleCpuComboBox.get()
		version = self.simpleVersion.get()
		ram = self.simpleRamComboBox.get()
		topology = self.simpleTopology.get()
		ethernetCount = int(self.simpleEthernetComboBox.get())
		spineCount = int(self.simpleSpineComboBox.get())
		leafCount = int(self.simpleLeafComboBox.get())
		spinePrefix = self.simpleSpinePrefix.get()
		leafPrefix = self.simpleLeafPrefix.get()
  
		if eq("", topology):
			messagebox.showwarning(title="warning", message="토폴로지이름을 입력해 주세요.")
			self.simpleTopologyTextBox.focus()
			return False
 
		if eq("", version):
			messagebox.showwarning(title="warning", message="버전을 입력해 주세요.")
			self.simpleVersionTextBox.focus()
			return False
 
		if eq("", spinePrefix):
			messagebox.showwarning(title="warning", message="spine prefix을 입력해 주세요.")
			self.simpleSpinePrefixTextBox.focus()
			return False

		if eq("", leafPrefix):
			messagebox.showwarning(title="warning", message="leaf prefix을 입력해 주세요.")
			self.simpleLeafPrefixTextBox.focus()
			return False
 
		if eq("", ethernetCount):
			messagebox.showwarning(title="warning", message="ethernet 갯수를 입력해 주세요.")
			self.simpleEthernetComboBox.focus()
			return False
    
    
		if (ethernetCount < leafCount):
			messagebox.showwarning(title="warning", message="ethernet 갯수가 leaf 갯수보다 작습니다.")
			return False

		if eq("", cpu):
			messagebox.showwarning(title="warning", message="cpu 갯수를 입력해 주세요.")
			self.simpleCpuComboBox.focus()
			return False
    
		if eq("", ram):
			messagebox.showwarning(title="warning", message="ram 사이즈를 입력해 주세요.")
			self.simpleRamComboBox.focus()
			return False
   
		self.gridReset()
  
		self.process.createSimpleTopology(topology=topology, spinePrefix=spinePrefix, leafPrefix=leafPrefix, spinesCount=spineCount, leafsCount=leafCount, cpu=cpu, ram=ram, version=version, ethernetCount=ethernetCount)
		self.treeview.insert('', END, text='', values=("간편 topology", topology + "이(가) 생성되었습니다.", datetime.now().strftime('%y-%m-%d %H:%M:%S')))
		
    

	def createConfig(self, session):
		self.nr = self.process.createConfig(session)
   
		self.gridReset()
   
		for host in self.nr.inventory.hosts:
    
			self.treeview.insert('', END, text='', values=(host, session + " config 생성 완료", datetime.now().strftime('%y-%m-%d %H:%M:%S')))

	def statusCheck(self, task: Task):

		result = Result(
			host=task.host,
			result=task.run(netmiko_send_command, command_string="show clock"),
		)
  
		return result

	def statusCheckCall(self):
   
		self.gridReset()
		self.nr = self.process.statusCheckCall()

		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "Connect... Failed"
			else:
				desc = "Connect... Done"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
