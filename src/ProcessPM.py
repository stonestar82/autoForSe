#-*- coding:utf-8 -*-
import os, re, json, shutil
from tqdm import tqdm
from datetime import datetime
from operator import eq, ne
from openpyxl import load_workbook
import paramiko, platform, sys
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

class ProcessPM():
	def __init__(self) -> None:
	
		if eq(platform.system().lower(), "windows"):
			self.path = "./pm"
			self.rootPath = "./"
		else:
			self.path = os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]) + "/pm"
			self.rootPath = os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]) + "/"
		self.hostnamePattern = "\nhostname\s(.*)"
		self.showtechPath = f"{self.path}/showtech/"
		self.showCmdPatternPrefix = "------------- "
		self.showCmdPatternSuffix = " -------------"
		self.showCmdPattern = f"(^{self.showCmdPatternPrefix}show .*?{self.showCmdPatternSuffix}$|^{self.showCmdPatternPrefix}.*?{self.showCmdPatternSuffix}$)"
		self.reportTemplatePath = f"{self.rootPath}inventory/pm/pmTemplate.xlsx"
		self.cliCmdChanged = {
			'show bfd neighbor detail': 'show bfd peers detail',
			'show env cooling': 'show system env cooling',
			'show env temperature': 'show system env temperature',
			'show interfaces': 'show interfaces all' ,
			'show interfaces status': 'show interfaces all status',
			'show environment power detail': 'show system environment power detail'
		}
  
		self.showTechFileGenList = ["show version detail", "show ip interface", "show system environment power detail", 
																"show system env temperature", "show system env cooling", "show processes top once", 
                                "dir recursive flash", "show interfaces all status", "show ntp status", "show logging",
                                "show running-config sanitized"]
  
	def showtechToReport(self):

		print("show tech-suport 파일 분석중\n\n")
		##### 폴더 생성ㅔㅛㅅ
		now = datetime.now()
		folderName = now.strftime("%Y%m%d")
		
  
		directory = f"{self.path}/report/{folderName}"
		tmpDirectory = f"{directory}/show/tmp"
		showDirectory = f"{directory}/show"
		dataDirectory = f"{directory}/data"
  
		if os.path.exists(directory) and os.path.exists(f"{directory}/정기점검_{folderName}.xlsx"):
			return False, "이미 작성된 정기점검 리포트가 있습니다."
  
		if not os.path.exists(directory):
			os.makedirs(directory)
   
		if not os.path.exists(showDirectory):
			os.makedirs(showDirectory)
  
		if not os.path.exists(dataDirectory):
			os.makedirs(dataDirectory)
  
  
		showtechList = os.listdir(self.showtechPath)
  
	
		if len(showtechList) == 0:
			return False, "show tech 파일이 없습니다."
		print("file list ", showtechList)
		for showtech in sorted(showtechList):
			print("\n\n", os.path.basename(showtech.strip()), " 처리중\n")
   
			if not os.path.exists(tmpDirectory):
				os.makedirs(tmpDirectory)
			print("show tech 파일 분석중", showtech)
			pass
			with open(f"{self.showtechPath}/{showtech}") as r:
				rl = r.readlines()
			r.close()
   
			config = []
			cmd = ""
			preCmd = ""
			
			p = re.compile(self.showCmdPattern)
			# bar = progressbar.ProgressBar(maxval=len(rl)).start()
			pbar = tqdm(rl)
			for line in pbar:
				cmdPattern = p.match(str(line).strip())
				
				# bar.update(idx)
				if cmdPattern:
					preCmd = cmd
					cmd = cmdPattern.group().replace(self.showCmdPatternPrefix, "").replace(self.showCmdPatternSuffix, "")
					cmd = re.sub(r"([/\|\/\:\\\'\^\<\>\"\*])", "", cmd)
		
					## eos 버전에 따라 변경된 cmd인경우 신규 cmd로 변경처리 
					if cmd in self.cliCmdChanged:
						cmd = self.cliCmdChanged[cmd]
						# print("changed cmd = ", cmd)
		
					if ne(preCmd, "") and preCmd in self.showTechFileGenList:
						# print(preCmd)
						with open(f"{tmpDirectory}/{preCmd}", "w") as f:
							f.write("".join(config))
							f.close

					config = []
		
				else:
					config.append(line)

			if cmd in self.showTechFileGenList:
				with open(f"{tmpDirectory}/{cmd}", "w") as f:
					f.write("".join(config))
				f.close()

			pbar.close()
    
			
   
			with open(f"{tmpDirectory}/show running-config sanitized") as r:
				c = r.read()
				result = re.findall(self.hostnamePattern, c)
				hostname = result[0]
				r.close()
   
   
			# time.sleep(0.2)
			if os.path.exists(f"{showDirectory}/{hostname}"):
				shutil.rmtree(f"{showDirectory}/{hostname}")
    
			os.rename(f"{tmpDirectory}", f"{showDirectory}/{hostname}")
   
		self.showToData(folderName)
  
		self.createReport(folderName)
  
		print("\n\n", "정기점검 리포트 작성완료\n")
  
		return True, ""
    
	def showToData(self, folder):
  
		print("\n\nshow tech-suport 데이터 처리중\n")
  
		directory = f"./pm/report/{folder}"
		showDirectory = f"{directory}/show"
		dataDirectory = f"{directory}/data"
  
		hostList = os.listdir(showDirectory)
  
		pbar = tqdm(hostList)
  
		for host in pbar:
			hostPath = f"{showDirectory}/{host}"
			# print(hostPath)
			# hostShowList = os.listdir(hostPath)
			# print(hostShowList)
   
			data = { "HOSTNAME-HOST": host}
   
			##### show version detail
			##### serial, model, total memory, free memory, free memory per, uptime, eos version
			with open(f"{hostPath}/show version detail") as f:
				
				c = f.read()
				f.seek(0)
				l = f.readlines()
			f.close()

			serial = str(re.findall("Serial number: (.*)", c)[0]).strip()
			model = re.findall("[A-Z0-9-]{8,30}", l[1])[0]
			totalMemory = re.findall("Total memory: (.*) kB", c)[0]
			freeMemory = re.findall("Free memory: (.*) kB", c)[0]
			memoryFree = int((int(totalMemory) - int(freeMemory)) / int(totalMemory) * 100)
			memoryFreePercent = f"{memoryFree}%"
			uptime = re.findall("Uptime: (.*)", c)[0]
			esoVersion = re.findall("Software image version: (.*)", c)[0]
			# memoryFreeStatus = "정상" if memoryFree < 70 else f"확인 필요 ( {memoryFreePercent} )"
			memoryFreeStatus = "정상" if memoryFree < 40 else f"확인 필요 ( {memoryFreePercent} )"
	
			data.setdefault("VERSION DETAIL-SERIAL", serial)
			data.setdefault("VERSION DETAIL-MODEL", model)
			data.setdefault("VERSION DETAIL-TOTAL MEMORY", str(totalMemory).strip())
			data.setdefault("VERSION DETAIL-FREE MEMORY", str(freeMemory).strip())
			data.setdefault("VERSION DETAIL-FREE MEMORY PERCENT", memoryFreePercent)
			data.setdefault("VERSION DETAIL-FREE MEMORY STATUS", memoryFreeStatus)
			data.setdefault("VERSION DETAIL-UPTIME", str(uptime).strip())
			data.setdefault("VERSION DETAIL-EOS VERSION", str(esoVersion).strip())
				
    
			# print(data)
    
    
			##### show ip interface brief
			##### host ip(mgmt)
			with open(f"{hostPath}/show ip interface") as f:
				c = f.read()
			f.close()
			## description 이 없는경우
			hostIp = re.findall("Management(\d) is up, line protocol is up \(connected\)\n  Internet address is (.*)", c)

			if hostIp:
				hostIp = list(hostIp[0])[1]

			## description 이 있는경우
			else:
     
				hostIp = re.findall("Management(\d) is up, line protocol is up \(connected\)\n(.*)\n  Internet address is (.*)", c)
    
				if hostIp:
					hostIp = list(hostIp[0])[2]
				else:
					hostIp = "N\A"
				
   
			# print(hostIp)
    
			data.setdefault("IP INTERFACE-HOSTIP", hostIp)


			##### show system venvironment power detail
			##### supply
			powerFind = False
   
			## 버전에 따라 다른듯
			powerDetail = f"{hostPath}/show system environment power detail"
   
			# if not os.path.exists(f"{hostPath}/show system environment power detail"):
			# 	powerDetail = f"{hostPath}/show environment power detail"
   
			with open(powerDetail) as f:
				l = f.readlines()	
			f.close()
   
			
			powerSupply = []
			psIdx = 0
			for r in l:
				if "PWR-" in r and not powerFind:
					powerFind = True

				if "PWR-" in r:
					psIdx = psIdx + 1
					if "Ok" in r:
						status = f"power-supply {psIdx} : 정상\n"
					elif "Loss" in r:
						status = f"power-supply {psIdx} : 파워케이블 결함\n"
					else:
						status = f"power-supply {psIdx} : 확인요망\n"

					powerSupply.append(status)
     
				if powerFind and not "PWR-" in r:
					break
   
			powerSupply[-1] = str(powerSupply[-1]).replace("\n", "")
			power = "".join(powerSupply)
			# print(power)

			data.setdefault("SYSTEM ENVIRONMENT POWER DETAIL-POWER SYSTEM", power)


			##### show system env temperature
			## status
			envTemperature = f"{hostPath}/show system env temperature"
   
			# if not os.path.exists(envTemperature):
			# 	envTemperature = f"{hostPath}/show env temperature"
    
			with open(envTemperature) as f:
				c = f.read()
			f.close()
			status = str(re.findall("System temperature status is:(.*)", c)[0]).strip()

			if eq("Ok", status):
				statusDetail = "정상" 
			elif eq("Overheating", status):
				statusDetail = "경고" 
			elif eq("Critical", status):
				statusDetail = "위험" 
			elif eq("Unknown", status):
				statusDetail = "스위치초기화" 
			elif eq("Sensor Failed", status):
				statusDetail = "온도센서 오작동" 
			else:
				statusDetail = "확인요망" 

			data.setdefault("SYSTEM ENV TEMPERATURE-STATUS", statusDetail)
   
   
			##### show system env coolling
			##### fan 정보
			envCooling = f"{hostPath}/show system env cooling"
			fanStatus = []
			# if not os.path.exists(envCooling):
			# 	envCooling = f"{hostPath}/show env cooling"
    
			with open(envCooling) as f:
				l = f.readlines()
			f.close()
			
			fanFirstPatter = re.compile("^1/1")
			fanPattern = re.compile("([0-9]{1,2}\/[0-9]{1,2}|PowerSupply[0-9]{1,2}/[0-9]{1,2}|PowerSupply[0-9]{1,2})")
			
			fanExsist = False
			for c in l:
				c = str(c).strip()
				if not fanExsist:
					fan = fanFirstPatter.match(c)
					if fan:
						fanExsist = True
				## 버전에 따라서 Stable Uptime이 별개의 줄로 되어 있는 경우가 있음
				#                              Config Actual                    Speed                   Stable
				# Fan             Status        Speed  Speed             Uptime Stability               Uptime

				# 															Config Actual                    Speed
				# Fan            Status  Speed  Speed             Uptime Stability
				
        #     Stable
        #     Uptime
				if fanExsist and eq("", c.replace("\n", "")):
					break
				elif fanExsist:
					statusDetail = "확인 요망"
					if " Ok " in c:
						statusDetail = "정상"
					elif " Failed " in c:
						statusDetail = "FAN 오작동"
					elif " Unknown " in c:
						statusDetail = "스위치 초기화"
					elif " Not Inserted " in c:
						statusDetail = "FAN 미장착"
					elif " Unsupported " in c:
						statusDetail = "소프트웨어 버전 미지원"
				
					r = "FAN " + str(fanPattern.match(c).group()) + " : " + statusDetail + "\n"
					fanStatus.append(r)
					# print(r)

			fanStatus[-1] = str(fanStatus[-1]).replace("\n", "")
			# print(fanStatus)
			data.setdefault("SYSTEM ENV COOLING-FAN STATUS", "".join(fanStatus))
					

			##### show processes top once
			#### CPU 사용량 및 가장높은 사용률 프로세스
			with open(f"{hostPath}/show processes top once") as f:
				c = f.read()
			f.close()
			## description 이 없는경우
			cpuRate = re.findall("\%Cpu\(s\):(.*)us", c)
			cpuRate = float(str(cpuRate[0]).strip())
   
			# if cpuRate < 50:
			##### 테스트를 위해 10으로 수정
			if cpuRate < 10:
				statusDetail = "정상"
			else:
				cpuRate = int(cpuRate)
				topProcess = str(re.findall("COMMAND\n(.*)", c)[0])
				topProcess = topProcess[topProcess.rfind(" ")+1:]
				statusDetail = f"확인 필요 ({cpuRate}% : {topProcess})"

			# print(hostIp)
    
			data.setdefault("ENV COOLING-FAN STATUS-CPU USE", statusDetail)
   
   
			##### dir /recursive flash:
			## flash 사용량
			with open(f"{hostPath}/dir recursive flash") as f:
				c = f.read()
			f.close()
			# print(c)
			## description 이 없는경우
			flash = list(re.findall("([0-9]*) bytes total \(([0-9]*) bytes free\)", c)[0])
			flashTotal = flash[0]
			flashFree = flash[1]
			flashUsed = (int(flashTotal) - int(flashFree))
			flashUsedPer = int(flashUsed / int(flashTotal) * 100)
   
			if flashUsedPer > 70:
				flashStatus = f"확인 필요 ( {flashUsedPer}% )"
			else:
				flashStatus = "정상"

			data.setdefault("DIR RECURSIVE FLASH-TOTAL", flashTotal)
			data.setdefault("DIR RECURSIVE FLASH-FREE", flashFree)
			data.setdefault("DIR RECURSIVE FLASH-USED", flashUsed)
			data.setdefault("DIR RECURSIVE FLASH-USED PERCENT", flashUsedPer)
			data.setdefault("DIR RECURSIVE FLASH-STATUS", flashStatus)
   
   
   
			##### show interfaces all status
			## 인터페이스 점검
			with open(f"{hostPath}/show interfaces all status") as f:
				l = f.readlines()
			f.close()
   
			etPattern = re.compile("Et[0-9]{1,2}\/[0-9]{1,2}\/[0-9]{1,2}|Et[0-9]{1,2}\/[0-9]{1,2}|Et[0-9]{1,2}")
			connectedPattern = re.compile(" connected | notconnect | disabled | errdisabled ")
   
			## errdisabled  show interfaces status errdisable
      ## speed-missconfig 찾아서 제외처리 GS3-HBB01_tech-support_2022-04-06.1004
			duplexPattern = re.compile(" full | half | a-full | auto | a-half ")
			totalPort = 0
			usePort = 0
			interfaces = {}
			halfPort = []
			for c in l:
				c = str(c).strip()
				
				interface = etPattern.match(c)
				if interface:
					connect = connectedPattern.search(c).group().strip()
					duplex = duplexPattern.search(c).group().strip()
					interface = interface.group().strip()
     
					interfaces.setdefault(interface, {
						"connected": connect,
						"duplex": duplex
					})
					# print(interface, connect, duplex)
					totalPort = totalPort + 1
     
					if eq("connected", connect):
						usePort = usePort + 1
      
					## half a-half 구분처리
					if eq("half", duplex) or eq("a-half", duplex):
						halfPort.append(interface)

			## 사용가능 포트
			useablePort = totalPort - usePort

			## 포트 사용률
			usePortPercent = int((usePort / totalPort) * 100)
   
			if usePortPercent < 70:
				portUseStatus = "정상"
			else:
				portUseStatus = f"확인 필요 ( {usePortPercent}% )"
    
			if len(halfPort) > 0:
				halfPort = ", ".join(halfPort) + " 확인필요"
			else:
				halfPort = "정상"
    
			totalStatus = f"사용률 70% 미만 = {portUseStatus}\nDuplex = {halfPort}"

			interfaces.setdefault("totalPort", totalPort)
			interfaces.setdefault("usePort", usePort)
			interfaces.setdefault("useablePort", useablePort)
   
			data.setdefault("INTERFACES ALL STATUS-INTERFACES", interfaces)
			data.setdefault("INTERFACES ALL STATUS-TOTAL PORTS", totalPort)
			data.setdefault("INTERFACES ALL STATUS-USE PORTS", usePort)
			data.setdefault("INTERFACES ALL STATUS-USEABLE PORTS", useablePort)
			data.setdefault("INTERFACES ALL STATUS-PORT USE STATUS", portUseStatus)
			data.setdefault("INTERFACES ALL STATUS-HALF PORTS", halfPort)
			data.setdefault("INTERFACES ALL STATUS-TOTAL STATUS", totalStatus)


			##### show ntp status
			## ntp 상태 체크
			with open(f"{hostPath}/show ntp status") as f:
				c = f.read()
			f.close()
   
			ntpPattern = re.compile("synchronised to NTP server|unsynchronised|NTP is disabled")
   
			ntp = ntpPattern.search(c).group()
			
			if eq("synchronised to NTP server", ntp):
				ntpStatus = "정상"
			elif eq("unsynchronised", ntp):
				ntpStatus = "NTP 설정 확인 필요"
			elif eq("NTP is disabled", ntp):
				ntpStatus = "NTP 사용 안함"
			else:
				ntpStatus = "확인 필요"
    
			data.setdefault("NTP STATUS-STATUS", ntpStatus)
   
			##### show loggin
			## buffer, console, monitor 로깅레벨 체크
			"""
			Buffer logging: level debugging
			Console logging: level informational
			Monitor logging: level informational
			"""
			logginLevel = "(emergencies|alerts|critical|errors|warnings|notifications|informational|debugging)"

			with open(f"{hostPath}/show logging") as f:
				c = f.read()
			f.close()

			buffLevel = re.findall(f"Buffer logging: level {logginLevel}", c)[0]
			consoleLevel = re.findall(f"Console logging: level {logginLevel}", c)[0]
			monitorLevel = re.findall(f"Monitor logging: level {logginLevel}", c)[0]
   
			logginLevelStatus = []
   
			if ne("informational", buffLevel) and ne("debugging", buffLevel):
				logginLevelStatus.append(f"Buffer logging level 확인 필요 ( {buffLevel} )")
    
			if ne("informational", consoleLevel) and ne("debugging", consoleLevel):
				logginLevelStatus.append(f"Console logging level 확인 필요 ( {consoleLevel} )")
      
			if ne("informational", monitorLevel) and ne("debugging", monitorLevel):
				logginLevelStatus.append(f"Monitor logging level 확인 필요 ( {monitorLevel} )")
    
			if len(logginLevelStatus) > 0:
				logginLevelStatus = "\n".join(logginLevelStatus)
			else:
				logginLevelStatus = "정상"
   
			data.setdefault("LOGGING-BUFFER LOGGING", buffLevel)
			data.setdefault("LOGGING-CONSOLE LOGGING", consoleLevel)
			data.setdefault("LOGGING-MONITOR LOGGING", monitorLevel)
			data.setdefault("LOGGING-LEVEL STATUS", logginLevelStatus)
   

			##### show hardware capacity
			## lem
			"""
			LEM으로 시작하지 않는경우
			2번째 탭이 공란인게 그룹 총합
			아래 내역이 카운팅되면(L2) 같이 올라감
			MAC                              Linecard0/0             8       0%       65528             0         65536          12
			MAC              L2              Linecard0/0             8 
   
			LEM으로 시작하고 MAC이 있음
			LEM               MAC                                   70       0%      786299             0        786432          93 
			"""

			with open(f"{dataDirectory}/{host}-data.json", "w", encoding='utf8') as f:   
				json.dump(data, f)
			f.close()
   
		pbar.close()
			# print(data)
   
	def createReport(self, folder):
		
		## template 엑셀 복사
		workbook = load_workbook(filename=self.reportTemplatePath, read_only=False, data_only=False)
  
		# print(workbook.sheetnames)

		directory = f"./pm/report/{folder}"
		showDirectory = f"{directory}/show"
		dataDirectory = f"{directory}/data"
  
		hostList = os.listdir(showDirectory)
  
		print("\n\n정기점검 리포트 작성중\n")
  
		pbar = tqdm(hostList)
		for host in pbar:
			hostData = f"{dataDirectory}/{host}-data.json"
			newSheet = workbook.copy_worksheet(workbook["template"])
			newSheet.title = host

			with open(hostData, "r", encoding='utf8') as f:
				data = json.load(f)
			f.close()

			# print(data)
			if newSheet.iter_rows():
				for row in newSheet.iter_rows():
					if ne("", row[7].value):
						if row[7].value in data:
							row[7].value = data[row[7].value]
    
		pbar.close()
  
		workbook.remove_sheet(workbook["template"])
		workbook.save(f"./pm/report/{folder}/정기점검_{folder}.xlsx")
		workbook.close()
    
    
		## template 시트 로드

		## 폴더내 각 호스트
  
		## 호스트 폴더내 data.json 로드
  
		## teplate에 대입 및 시트 생성
  

		## template 시트 삭제
		return False


	def getShowTechFile(self, hosts):

		"""
		D1-BL-02(config)#bash 
		[admin@D1-BL-02 tech-support]$ ls -tr | tail -1 
		D1-BL-02_tech-support_2023-02-10.0937.log.gz
  	"""

		print(hosts)
  
		client = paramiko.SSHClient()
		client.load_system_host_keys()
		client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
		
  
		# for host in hosts:
			# client.connect(server, port, user, password)

		return False


	def showtechSummary(self, analysisPath):
		
		df = pd.read_excel(analysisPath, sheet_name="System", engine='openpyxl').fillna(0)
  
		less20 = df[(df['Memory %'] >= 0) & (df['Memory %'] < 20)]["Memory %"].count()
  
		less40 = df[(df['Memory %'] >= 20) & (df['Memory %'] < 40)]["Memory %"].count()
  
		less60 = df[(df['Memory %'] >= 40) & (df['Memory %'] < 60)]["Memory %"].count()
  
		less80 = df[(df['Memory %'] >= 40) & (df['Memory %'] < 80)]["Memory %"].count()
  
		less100 = df[(df['Memory %'] >= 80) & (df['Memory %'] <= 100)]["Memory %"].count()
  
		total = less20 + less40 + less60 + less80 + less100
  
		frequency = [less20 / total]
  
		labels = ['0~20 %']
		colors = ['whitesmoke']
  
		if (less40 > 0):
			frequency.append(less40 / total)
			colors.append('#ff9999')
			labels.append('20~40 %')
		if (less60 > 0):
			frequency.append(less60 / total)
			colors.append('#ffc000')
			labels.append('40~60 %')
		if (less80 > 0):
			frequency.append(less80 / total)
			colors.append('#8fd9b6')
			labels.append('60~80 %')
		if (less100 > 0):
			frequency.append(less100 / total)
			colors.append('#d395d0')
			labels.append('80~100 %')
  

		labels_frequency = zip(labels,frequency,colors) 
		labels_frequency = sorted(labels_frequency,key=lambda x: x[1],reverse=True)
		
		sorted_labels = [x[0] for x in labels_frequency] ## 정렬된 라벨
		sorted_frequency = [x[1] for x in labels_frequency] ## 정렬된 빈도수
		sorted_colors = [x[2] for x in labels_frequency] ## 정렬된 색상
		
		fig = plt.figure(figsize=(8,8)) ## 캔버스 생성
		fig.set_facecolor('white') ## 캔버스 배경색을 하얀색으로 설정
		ax = fig.add_subplot() ## 프레임 생성
		
		pie = ax.pie(sorted_frequency, ## 파이차트 출력
					startangle=90, ## 시작점을 90도(degree)로 지정
					counterclock=False, ## 시계방향으로 그려짐
					colors = sorted_colors, ## 색상 지정
					)
		
		total = np.sum(frequency) ## 빈도수 합
		
		threshold = 5
		sum_pct = 0 ## 퍼센티지
		count_less_5pct = 0 ## 5%보다 작은 라벨의 개수
		spacing = 0.1
		for i,l in enumerate(sorted_labels):
				ang1, ang2 = ax.patches[i].theta1, ax.patches[i].theta2 ## 파이의 시작 각도와 끝 각도
				center, r = ax.patches[i].center, ax.patches[i].r ## 파이의 중심 좌표
				
				## 비율 상한선보다 작은 것들은 계단형태로 만든다.
				if sorted_frequency[i]/total*100 < threshold:
						x = (r/2+spacing*count_less_5pct)*np.cos(np.pi/180*((ang1+ang2)/2)) + center[0] ## 텍스트 x좌표
						y = (r/2+spacing*count_less_5pct)*np.sin(np.pi/180*((ang1+ang2)/2)) + center[1] ## 텍스트 y좌표
						count_less_5pct += 1
				else:
						x = (r/2)*np.cos(np.pi/180*((ang1+ang2)/2)) + center[0] ## 텍스트 x좌표
						y = (r/2)*np.sin(np.pi/180*((ang1+ang2)/2)) + center[1] ## 텍스트 y좌표
				
				## 퍼센티지 출력
				if i < len(labels) - 1:
						sum_pct += float(f'{sorted_frequency[i]/total*100:.2f}')
						ax.text(x,y,f'{sorted_frequency[i]/total*100:.2f}%',ha='center',va='center',fontsize=12)
				else: ## 마지막 파이 조각은 퍼센티지의 합이 100이 되도록 비율을 조절
						ax.text(x,y,f'{100-sum_pct:.2f}%',ha='center',va='center',fontsize=12)
		
		plt.legend(pie[0],sorted_labels) ## 범례
  
		plt.savefig("./pm/report/pie.png")
	
		# work = pd.read_excel(analysisPath, engine='openpyxl')
		# writer = pd.ExcelWriter(analysisPath, engine = 'openpyxl')
		# work.to_excel(writer)

		# workSheet = writer.sheets["Chart"]
		# workSheet.insert_image("C2", "pie.png")
		# writer.save()