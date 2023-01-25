import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import *
import yaml, json
import asyncio, os
from nornir import InitNornir
from nornir.plugins import *
from nornir.plugins.inventory import *
from nornir.plugins.runners import *
from nornir.plugins.functions import *
from nornir.plugins.processors import *
from nornir.plugins.tasks import *
from nornir.plugins.connections import *
from nornir_netmiko.tasks import netmiko_send_config, netmiko_send_command
from nornir.core.task import Task, Result
from jinja2 import Template
from datetime import datetime
from openpyxl import load_workbook
from generators.BlankNone import *
from generators.generateInventory import generateInventory 
from operator import eq
import zipfile
from src.ProcessImpl import ProcessImpl

class Processing():
	def __init__(self, asyncio=asyncio, process=ProcessImpl):
		self.asyncio = asyncio
		self.process = process
		self.loop = self.asyncio.get_event_loop()
		self.root = tk.Tk()
		self.root.title("GlobalTelecom & i-Cloud")
		self.root.geometry("640x500+600+100") ## w, h, x, y
		self.root.resizable(False, False)
  
		self.db = "./db/db.xlsx"
		self.inventory = "./inventory.xlsx"
		self.lastConfigGen = "" ## 마지막으로 실행한 config 생성 session
		self.fullSession = {"fullv4":["init", "base", "loop0", "p2pip", "bgpv4", "etcport", "vxlan"] , "fullv6": ["init", "base", "loop0", "p2pipv6", "bgpv6", "etcport", "vxlan"]}
		self.sessions = list(set(self.fullSession["fullv4"] + self.fullSession["fullv6"]))
  
		self.ymlInit()
		self.norInit()
		##### grid S #####
		self.frame_top = ttk.Frame(self.root)
		self.frame_top.pack(side="top")


		self.treeview=ttk.Treeview(self.frame_top, columns=["one", "two", "three"])
		self.treeview.tag_configure("green",foreground='green')
		self.treeview.tag_configure("red",foreground='red')

		self.treeview.column("#1", width=180, anchor="center")
		self.treeview.heading("one", text="name", anchor="center")

		self.treeview.column("#2", width=330, anchor="w")
		self.treeview.heading("two", text="description", anchor="center")
  
		self.treeview.column("#3", width=120, anchor="w")
		self.treeview.heading("three", text="datetime", anchor="center")


		self.treeview["show"] = "headings"
		self.treeview.pack()
		##### grid E #####

		##### 메모 S #####
		self.frame_middle = ttk.Frame(self.root)
		self.frame_middle.pack(side="top")
  
		self.iddLabel = ttk.Label(self.frame_middle, text="메모")
		self.iddLabel.grid(column=0, row=1)
		self.memo = StringVar()
		self.memoTextbox = ttk.Entry(self.frame_middle, width=50, textvariable=self.memo)
		self.memoTextbox.grid(column=1 , row=1)
		self.memoTextbox.focus()  
		##### 메모 E #####
	  
		##### 백업 및 grid 선택 배포 S #####
		self.frameLeft = ttk.Frame(self.root)
		self.frameLeft.pack(side="left")
  
		self.buttonStatus = tk.Button(self.frameLeft, text="상태체크", width=18, command=lambda: self.loop.create_task(self.statusCheckCall()))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameLeft, text="Config 백업", width=18, command=lambda: self.loop.create_task(self.backupConfigCall()))
		self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.frameLeft, text="Config 백업 List", width=18, command=lambda: self.loop.create_task(self.getBackupConfigList()))
		self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.frameLeft, text="Config 백업 선택 배포", width=18, command=lambda: self.loop.create_task(self.selectedConfigCall()))
		self.buttonSelectedDeploy.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### 백업 및 grid 선택 배포 E #####
  
  
		##### session 별 config 생성 및 배포 S #####
		self.frameMiddle = ttk.Frame(self.root)
		self.frameMiddle.pack(side="left")
  
		self.buttonStatus = tk.Button(self.frameMiddle, text="init Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("init")))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameMiddle, text="base Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("base")))
		self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonList = tk.Button(self.frameMiddle, text="loop0 Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("loop0")))
		self.buttonList.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.frameMiddle, text="etcport Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("etcport")))
		self.buttonList.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonList = tk.Button(self.frameMiddle, text="vxlan Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("vxlan")))
		self.buttonList.grid(row=4, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonStatus = tk.Button(self.frameMiddle, text="p2p IPv4 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("p2pip")))
		self.buttonStatus.grid(row=0, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameMiddle, text="BGP IPv4 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("bgpv4")))
		self.buttonBackUp.grid(row=1, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonStatus = tk.Button(self.frameMiddle, text="p2p IPv6 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("p2pipv6")))
		self.buttonStatus.grid(row=2, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameMiddle, text="BGP IPv6 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("bgpv6")))
		self.buttonBackUp.grid(row=3, column=2, sticky=tk.W, padx=(8, 8), pady=(8, 8))


		self.buttonSelectedDeploy = tk.Button(self.frameMiddle, text="Config 배포", width=40, command=lambda: self.loop.create_task(self.sendConfigCall()))
		self.buttonSelectedDeploy.grid(row=5, columnspan=3, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### session 별 config 생성 및 배포 E #####

		##### full config 생성 및 배포 S #####
		self.frameRight = ttk.Frame(self.root)
		self.frameRight.pack(side="left")
  
		self.buttonStatus = tk.Button(self.frameRight, text="full IPv4 Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("fullv4")))
		self.buttonStatus.grid(row=0, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameRight, text="full IPv6 Config 생성", width=18, command=lambda: self.loop.create_task(self.createConfig("fullv6")))
		self.buttonBackUp.grid(row=1, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
  
		self.buttonBackUp = tk.Button(self.frameRight, text="토폴로지 생성", width=18, command=lambda: self.loop.create_task(self.createDefaultTopology()))
		self.buttonBackUp.grid(row=2, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))

		self.buttonSelectedDeploy = tk.Button(self.frameRight, text="※※※ 초기화 ※※※", width=18, command=lambda: self.loop.create_task(self.cleanConfigCall()))
		self.buttonSelectedDeploy.grid(row=3, column=1, sticky=tk.W, padx=(8, 8), pady=(8, 8))
		##### full config 생성 및 배포 S #####
  
		self.root.protocol("WM_DELETE_WINDOW", self.windowButtonClose)

  
		# self.proccessingState = False
		# self.root.update()
  
	def ymlInit(self):
		file_location = self.inventory

	
		with open("./excelEnvriment.json", "r") as f:
			excelVar = json.load(f)
			f.close()

		workbook = load_workbook(filename=file_location, read_only=False, data_only=True)
		fabric_name = getExcelSheetValue(workbook, excelVar["all"]["fabricName"])

		avd = {
			"inventory": None,
			"group_vars": {
				fabric_name: None
			}
		}

		avd["inventory"] = generateInventory(file_location, excelVar)

		with BlankNone(), open("./inventory/inventory.yml", "w") as inv:
			inv.write(yaml.dump(avd["inventory"], sort_keys=False))
  
	def norInit(self):
		self.nr = InitNornir(config_file="./config.yml")
  
	def windowButtonClose(self):
		# print("close!!")
		try:
			self.loop.complete()
			self.loop.stop()
			self.loop.close()
		except:
			print("close!")		
		finally:
			self.root.quit()
			self.root.destroy()
			print("finally close")
   
	def cleanConfig(self, task: Task):
		mgmtVrf = task.host.defaults.data["MGMT_VRF"]
		mgmtGw = task.host.defaults.data["MGMT_GW"]
		mgmtInterface = task.host.defaults.data["MGMT_INTERFACE"]
  
		vrfConfig = f"vrf {mgmtVrf}"
		vrfInstance = f"vrf instance {mgmtVrf}"


		cleanConfig = ["configure session", \
									"rollback clean-config", \
									f"{vrfInstance}", \
									f"interface {mgmtInterface}", \
									f"{vrfConfig}", \
									f"ip address {task.host.hostname}/24", \
									f"username {task.host.username} secret {task.host.password} privilege 15", \
									f"ip route {vrfConfig} 0/0 {mgmtGw}", \
									"ip routing", \
									"logging buffered 1000", \
									"logging console informational", \
									"logging monitor informational", \
									"logging synchronous level all", \
									"commit"]
  
		result = task.run(netmiko_send_config, config_commands=cleanConfig)
  
		self.norInit()
  
	def cleanConfigCall(self):
		self.gridReset()
		self.nr.run(task=self.cleanConfig)
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "초기화 실패"
			else:
				desc = "초기화 완료"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
  
	async def getBackupConfigList(self):
		self.gridReset()
  
		workbook = load_workbook(filename=self.db, read_only=True, data_only=True)
		sheet = workbook["db"]

		# print(sheet.max_row)
		if sheet.iter_rows():
			self.lastConfigGen = "backup"
			for row in sheet.iter_rows():
				self.treeview.insert('', END, text='', values=(row[1].value, row[0].value, row[2].value))
		else:
			messagebox.showwarning(title="info", message="백업내역이 없습니다.")

	async def selectedConfigCall(self):

		if self.lastConfigGen != "backup":
			messagebox.showwarning(title="warning", message="Config 백업 List 호출후 사용하실 수 있습니다.")
			return False
   
		iid = self.treeview.focus()
		if iid:
			item = self.treeview.item(iid).get("values")[0]
			desc = self.treeview.item(iid).get("values")[1]
		else:
			messagebox.showwarning(title="warning", message="선택된 영역이 없습니다.")
			return FALSE
 
		directory = f"./inventory/config_backup/{item}"
		# print(f"config_backup/{item}")
		if not os.path.exists(directory):
			messagebox.showwarning(title="warning", message="백업내역이 없습니다. 확인후 다시 시도하세요.")
			return FALSE
 
		
		self.gridReset()
		self.nr.run(task=self.sendConfig, dir=f"config_backup/{item}")
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				descDetail = desc + " config 배포 실패"
			else:
				descDetail = desc + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, descDetail, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.norInit()
		
		# print("done!!")
			
	def sendConfig(self, task: Task, dir):
		"""
		config 배포
		"""
		result = Result(
			host=task.host,
			result=task.run(netmiko_send_config, config_file=f"./inventory/{dir}/{task.host.name}.cfg")
		)
		# print("print host-----------")
		# print(result.result[0].result)
		return result

	async def sendConfigCall(self):
		"""
		config 배포 호출
		"""

		if self.lastConfigGen == "backup":
			messagebox.showwarning(title="warning", message="config 생성 후 사용하실 수 있습니다.")
			return False
 
		self.gridReset()
		self.nr.run(task=self.sendConfig, dir="config")
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = self.lastConfigGen + " config 배포 실패"
			else:
				desc = self.lastConfigGen + " config 배포 완료"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
   
		self.norInit()
  
	def gridReset(self):
		"""
		Grid 내용 초기화
		"""
		x = self.treeview.get_children()
		for item in x:
			self.treeview.delete(item)
	
		self.nr.data.reset_failed_hosts()
   
	async def show(self):
		
		while True:
			
			self.root.update()
					
			await asyncio.sleep(.01)
   
	def backupConfig(self, task: Task, now):
		
		taskHost = task.host
		taskResult = task.run(netmiko_send_command, command_string="show running-config")

		with open(f"./inventory/config_backup/{now}/{taskHost}.cfg", "w") as inv:
			inv.write(taskResult[0].result)
			inv.close()
				
		return Result(
			host=taskHost,
			result=taskResult,
		)

	async def backupConfigCall(self):
		self.gridReset()
		  
		now = datetime.now()
		folderName = now.strftime("%y%m%d%H%M%S")
		nowDate = now.strftime('%y-%m-%d %H:%M:%S')
  
		directory = f"./inventory/config_backup/{folderName}"
		if not os.path.exists(directory):
			os.makedirs(directory)
  
		self.nr.run(task=self.backupConfig, now=folderName)
  
		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "Config 백업 실패"
			else:
				desc = "Config 백업 완료"

			self.treeview.insert('', END, text='', values=(host, desc, nowDate))
   
		memo = self.memo.get()
		self.memo.set("")
  
		if memo == "":
			memo = "config 백업"
   
		workbook = load_workbook(filename=self.db, read_only=False, data_only=True)
		sheet = workbook["db"]
		sheet.append([memo, folderName, nowDate])
		workbook.save(self.db)

	async def createTopology(self):

		spineSwitchs = {}
		leafSwitchs = {}

		with open("./inventory/hosts.yml") as f:
			hosts = yaml.load(f, Loader=yaml.FullLoader)
			f.close()

		for host in hosts:
			hostType = hosts[host]["data"]["TYPE"]

			if (hostType.upper() == "SPINE"):
				spineSwitchs.setdefault(
					host, {}
				)
			else:
				leafSwitchs.setdefault(
					host, {}
				)
    
		spinesCount = len(spineSwitchs)
		leafsCount = len(leafSwitchs)
		width = 1920
		space = 200 ## switch icon 사이 간격
		spineTopMargin = 200
		leafTopMargin = 600
		iconWidth = 50
  
		spineIdx = 0
		leafIdx = 0
  
		## spine icon 시작 위치값
		s_start = int(width / spinesCount - space / spinesCount)
		if eq(1, spinesCount):
			s_start = int(width / 2 - (iconWidth / 2))
  
		## leaf icon 시작 위치값
		l_start = int(width / leafsCount + space - (iconWidth/2))
  
		for host in hosts:
			hostType = hosts[host]["data"]["TYPE"]
			# print("host type = ", host)
			if (hostType.upper() == "SPINE"):
				spineIdx = spineIdx + 1
				spineSwitchs[host]= {
															"ID": spineIdx,
															"LEFT": s_start + (spineIdx * space),
															"TOP": spineTopMargin
														}
				
			else:
				leafIdx = leafIdx + 1
				leafSwitchs[host]= {
															"ID": leafIdx + spinesCount,
															"LEFT": l_start + (leafIdx * space),
															"TOP": leafTopMargin
														}
    
  
		with open("./inventory/topologyInterfaces.yml") as f:
			topologyInterfaces = yaml.load(f, Loader=yaml.FullLoader)
			f.close()

		topologyName = "TESTTOPOLOGY"
		data = {
			"NAME": topologyName,
			"EOS_VERSION": "veos-4.28.5M",
			"SPINES": spineSwitchs,
			"LEAFS": leafSwitchs,
			"INTERFACES": topologyInterfaces
		}
		
		with open('./inventory/templates/topology/topology.j2') as f:
			template = Template(f.read())
		with BlankNone(), open(f'./{topologyName}.unl', "w", encoding='utf8') as reqs:   
				reqs.write(template.render(**data))
				reqs.close() 
    
    
		zipOutput = f'./{topologyName}.zip'
		file = f'./{topologyName}.unl'
		zipFile = zipfile.ZipFile(zipOutput, "w")
		
		zipFile.write(file, compress_type=zipfile.ZIP_DEFLATED)

		zipFile.close()
  
		os.remove(file)
  
  
	async def createDefaultTopology(self):

		spineSwitchs = {}
		leafSwitchs = {}
  
		spinesCount = 4
		leafsCount = 16
		width = 1920
		space = 200 ## switch icon 사이 간격
		spineTopMargin = 200
		leafTopMargin = 600
		iconWidth = 50
  
		spineIdx = 0
		leafIdx = 0
  
		## spine icon 시작 위치값
		# s_start = int(width / spinesCount + space - (iconWidth/2))
		s_start = int((width/2) - (((spinesCount /2) * iconWidth) + (space / 2) + (space * (spinesCount /2 - 1))))
		# s_start = int(width - (((spinesCount /2) * iconWidth) + (space / 2) + (space * (spinesCount /2 - 1))))
		# print(s_start)
		if eq(1, spinesCount):
			s_start = int(width / 2 - (iconWidth / 2))
   
		if s_start < 0:
			s_start = 10
  
		## leaf icon 시작 위치값
		# l_start = int(width / leafsCount + space - (iconWidth/2))
		l_start = int((width/2) - (((leafsCount /2) * iconWidth) + (space / 2) + (space * (leafsCount /2 - 1))))
		if l_start < 0:
			l_start = 10
     
		for i in range(1, spinesCount + 1):
			spineIdx = spineIdx + 1
			spineName = "Spine-0" + str(i)
			spineSwitchs.setdefault(
					spineName, {
						"ID": spineIdx,
						"LEFT": s_start + (spineIdx * space),
						"TOP": spineTopMargin
					}
			)
   
		space = space - (leafsCount * 3)
		if space < 50:
			space = 50
   
		for i in range(1, leafsCount + 1):
			leafIdx = leafIdx + 1
			leafName = "Leaf-0" + str(i)
			leafSwitchs.setdefault(
					leafName, {
						"ID": leafIdx + spinesCount,
						"LEFT": l_start + (leafIdx * space),
						"TOP": leafTopMargin
					}
			)

		topologyInterfaces = {}
  
		id = 0
		
		leafPort = 0
		for spine in spineSwitchs:
			leafPort = leafPort + 1
			spinePort = 0
			for leaf in leafSwitchs:
				id = id + 1
				spinePort = spinePort + 1
    
				topologyInterfaces.setdefault(
					id, {
						"START": spine,
						"SPORT": "Et" + str(spinePort),
						"END": leaf,
						"EPORT": "Et" + str(leafPort),
					}
				)
  
		
		topologyName = "TESTTOPOLOGY1"
		data = {
			"NAME": topologyName,
			"EOS_VERSION": "veos-4.28.5M",
			"SPINES": spineSwitchs,
			"LEAFS": leafSwitchs,
			"INTERFACES": topologyInterfaces
		}
		
		# print(data)
		with open('./inventory/templates/topology/topology.j2') as f:
			template = Template(f.read())
		with BlankNone(), open(f'./{topologyName}.unl', "w", encoding='utf8') as reqs:   
				reqs.write(template.render(**data))
				reqs.close() 
    
    
		zipOutput = f'./{topologyName}.zip'
		file = f'./{topologyName}.unl'
		zipFile = zipfile.ZipFile(zipOutput, "w")
		
		zipFile.write(file, compress_type=zipfile.ZIP_DEFLATED)

		zipFile.close()
  
		os.remove(file)
    

	async def createConfig(self, session):
		self.lastConfigGen = session
  
		workbook = load_workbook(filename=self.inventory, read_only=True, data_only=True)
  
		j2 = ""
  
		if session in self.fullSession:
			for ss in self.fullSession[session]:
				sheet = workbook[ss]
				if sheet.iter_rows():
					for row in sheet.iter_rows():
						j2 = j2 + row[0].value + "\n"
		else:
			sheet = workbook[session]
			if sheet.iter_rows():
				for row in sheet.iter_rows():
					j2 = j2 + row[0].value + "\n"
				

		template = Template(j2)
   
		self.gridReset()
   
		for host in self.nr.inventory.hosts:
			data = {}
			for k in self.nr.inventory.hosts[host].keys():
				data.setdefault(k, self.nr.inventory.hosts[host][k])

			with open("./inventory/config/" + host + ".cfg", "w", encoding='utf8') as reqs:   
				reqs.write(template.render(**data))
				reqs.close() 
    
			self.treeview.insert('', END, text='', values=(host, session + " config 생성 완료", datetime.now().strftime('%y-%m-%d %H:%M:%S')))

	def statusCheck(self, task: Task):

		result = Result(
			host=task.host,
			result=task.run(netmiko_send_command, command_string="show clock"),
		)
  
		return result

	async def statusCheckCall(self):
   
		self.gridReset()
		self.nr.data.reset_failed_hosts()
  
		result = self.nr.run(task=self.statusCheck)

		for host in self.nr.inventory.hosts:
			if host in self.nr.data.failed_hosts:
				desc = "Connect... Failed"
			else:
				desc = "Connect... Done"

			self.treeview.insert('', END, text='', values=(host, desc, datetime.now().strftime('%y-%m-%d %H:%M:%S')))
